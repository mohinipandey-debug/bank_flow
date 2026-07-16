import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import datetime
import io
from database import (
    get_all_transactions, get_summary, get_uncategorized,
    get_large_debits, get_missing_statements, reload_categories, init_db,
    get_cashflow, manual_categorize, get_all_categories,
    revert_manual_category, get_category_audit,
    get_investment_register, update_investment_opening,
    add_investment_scheme, delete_investment_scheme,
    get_investment_summary, get_available_months,
    get_transaction_count, get_closing_balance, get_paginated_transactions,
    get_transfer_reconciliation, get_monthly_trend,
    get_yoy_comparison, get_top_expenses_comparison, get_weekly_cash_position,
    log_upload, get_upload_trail, delete_upload,
    get_account_balances, get_cash_at_stores, set_cash_at_stores,
    get_investment_transactions, tag_investment_transaction,
    get_investment_movements,
    add_manual_investment, get_manual_investments,
    delete_manual_investment, get_manual_investment_totals,
    get_entity_closing_balance, get_summary_with_bank,
    import_investment_excel, get_investment_kpis,
    get_entity_investment_balances,
)
from config import ENTITIES, LARGE_DEBIT_THRESHOLD, DATABASE_FILE

# ── Financial-year constants ───────────────────────────────────────────────────
FY_OPTIONS = ["FY2425", "FY2526", "FY2627"]
FY_LABELS  = {
    "FY2425": "FY 2024-25",
    "FY2526": "FY 2025-26",
    "FY2627": "FY 2026-27",
}
FY_BOUNDS  = {
    "FY2425": (datetime.date(2024, 4, 1), datetime.date(2025, 3, 31)),
    "FY2526": (datetime.date(2025, 4, 1), datetime.date(2026, 3, 31)),
    "FY2627": (datetime.date(2026, 4, 1), datetime.date(2027, 3, 31)),
}
from utils.formatters import fmt_inr, _inr, _td, _td_dash

# ── Cached DB wrappers ────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def _cached_summary(entity, month, financial_year):
    return get_summary(
        entity=entity if entity != "All" else None,
        month=month   if month  != "All" else None,
        financial_year=financial_year,
    )

@st.cache_data(ttl=600)
def _cached_months(financial_year):
    return get_available_months(financial_year=financial_year)

@st.cache_data(ttl=300)
def _cached_uncategorized(entity, bank, month, financial_year):
    return get_uncategorized(
        entity=entity             if entity         != "All" else None,
        bank=bank                 if bank           != "All" else None,
        month=month               if month          != "All" else None,
        financial_year=financial_year if financial_year and financial_year != "All" else None,
    )

@st.cache_data(ttl=300)
def _cached_summary_full(entity, bank, month, date_from, date_to):
    return get_summary_with_bank(
        entity=entity   if entity != "All" else None,
        bank=bank       if bank   != "All" else None,
        month=month     if month  != "All" else None,
        date_from=date_from if date_from else None,
        date_to=date_to     if date_to   else None,
    )

@st.cache_data(ttl=60)
def _cached_entity_balance(entity):
    return get_entity_closing_balance(entity)

@st.cache_data(ttl=300)
def _cached_count(entity, bank, month, financial_year):
    return get_transaction_count(
        entity=entity if entity != "All" else None,
        bank=bank     if bank   != "All" else None,
        month=month   if month  != "All" else None,
        financial_year=financial_year,
    )

@st.cache_data(ttl=300)
def _cached_closing(entity, bank, month, financial_year):
    return get_closing_balance(
        entity=entity if entity != "All" else None,
        bank=bank     if bank   != "All" else None,
        month=month   if month  != "All" else None,
        financial_year=financial_year,
    )

@st.cache_data(ttl=300)
def _cached_inv_header_totals():
    """Lightweight FD+MF totals for header band — avoids full investment query on every tab."""
    _d = get_investment_summary()
    return {
        "fd": sum(r["current_value"] for r in _d if r["scheme_type"] == "FD"),
        "mf": sum(r["current_value"] for r in _d if r["scheme_type"] == "MF"),
    }

@st.cache_data(ttl=60)
def _cached_inv_bals():
    return get_entity_investment_balances()
from queries.cashflow_queries import fetch_cf
from tabs.overview import render_overview
from tabs.review_queue import render_review_queue

# ── Live-view badge — shown in tabs that are NOT filtered by FY ───────────────
LIVE_BANNER = """
<div style="display:inline-flex; align-items:center; gap:8px;
     background:#F0FDF4; border:1px solid #86EFAC;
     border-radius:20px; padding:4px 14px; margin-bottom:16px;
     font-size:12px; color:#166534; font-weight:500;">
    <span style="width:8px; height:8px; background:#22C55E;
          border-radius:50%; display:inline-block;
          animation:pulse 2s infinite;"></span>
    Live &middot; All periods &middot; All entities
</div>
<style>
@keyframes pulse {
    0%, 100% { opacity:1; }
    50%       { opacity:0.4; }
}
</style>
"""


def fmt_cr(x):
    """Format number in Crores with 2 decimals. e.g. 52700000 → ₹5.27 Cr"""
    try:
        x = float(x)
        if x == 0:
            return "₹0"
        sign = "-" if x < 0 else ""
        x = abs(x)
        if x >= 1e7:
            return f"{sign}₹{x/1e7:.2f} Cr"
        elif x >= 1e5:
            return f"{sign}₹{x/1e5:.2f} L"
        else:
            return f"{sign}₹{x:,.0f}"
    except Exception:
        return str(x)


def fmt_cf(x):
    """Cash Flow formatter — negative as (₹X,XX,XXX), positive as ₹X,XX,XXX."""
    try:
        x = float(x)
        if x == 0:
            return "—"
        if x < 0:
            return f"(₹{abs(x):,.0f})"
        return f"₹{x:,.0f}"
    except Exception:
        return str(x) if x else "—"


st.set_page_config(
    page_title="BankFlow | Ventures & Stores",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# TO ADD YOUR COMPANY LOGO:
# 1. Save your logo file as 'logo.png' in the BankFlow folder
#    (same folder as dashboard.py)
# 2. Replace the logo placeholder block in the sidebar with:
#
#    from PIL import Image
#    logo = Image.open("logo.png")
#    st.sidebar.image(logo, use_container_width=True)
#
# 3. pip install Pillow  (run once in terminal)
# 4. Restart Streamlit

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

/* ═══════════════════════════════════════════════════════════
   DESIGN TOKENS
═══════════════════════════════════════════════════════════ */
:root {
  /* Surfaces */
  --bg-app:        #E8ECF3;
  --bg-surface:    #FFFFFF;
  --bg-raised:     #F4F6FA;
  --bg-inset:      #F0F3F8;

  /* Borders */
  --border-xs:     #EDF0F5;
  --border-sm:     #E2E7EF;
  --border-md:     #C8D3E0;

  /* Text */
  --tx-1:          #0A1628;
  --tx-2:          #3D4F66;
  --tx-3:          #6B7A90;
  --tx-4:          #96A3B4;

  /* Brand */
  --navy-900:      #081528;
  --navy-800:      #0F2044;
  --navy-700:      #1B2B4B;
  --navy-600:      #243654;
  --navy-500:      #2E4A72;
  --blue-600:      #1E40AF;
  --blue-500:      #2563EB;
  --blue-400:      #3B82F6;
  --blue-100:      #DBEAFE;
  --blue-50:       #EFF6FF;

  /* Semantic */
  --green-700:     #15803D;
  --green-500:     #22C55E;
  --green-400:     #4ADE80;
  --green-50:      #F0FDF4;
  --green-border:  #BBF7D0;

  --red-700:       #B91C1C;
  --red-500:       #EF4444;
  --red-400:       #F87171;
  --red-50:        #FFF1F2;
  --red-border:    #FECDD3;

  --amber-700:     #B45309;
  --amber-50:      #FFFBEB;
  --amber-border:  #FDE68A;

  /* Shadows — three tiers */
  --shadow-xs:     0 1px 2px rgba(10,22,40,0.04);
  --shadow-sm:     0 1px 3px rgba(10,22,40,0.06), 0 2px 8px rgba(10,22,40,0.04);
  --shadow-md:     0 2px 6px rgba(10,22,40,0.07), 0 4px 16px rgba(10,22,40,0.05);
  --shadow-lg:     0 4px 12px rgba(10,22,40,0.09), 0 8px 28px rgba(10,22,40,0.06);
  --shadow-inset:  inset 0 1px 2px rgba(10,22,40,0.05);

  /* Radius */
  --r-sm:  6px;
  --r-md:  8px;
  --r-lg:  12px;
  --r-xl:  16px;

  /* Transitions */
  --t-fast:  all 0.12s ease;
  --t-base:  all 0.18s ease;
}

/* ═══════════════════════════════════════════════════════════
   BASE RESET
═══════════════════════════════════════════════════════════ */
html, body, [class*="css"] {
  font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
  -webkit-font-smoothing: antialiased !important;
  text-rendering: optimizeLegibility !important;
}

.block-container {
  padding-top: 0 !important;
  padding-bottom: 1.5rem !important;
  max-width: 100% !important;
}

.stApp {
  background: var(--bg-app) !important;
}

/* Custom scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: var(--border-md); border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: var(--tx-4); }

/* ═══════════════════════════════════════════════════════════
   SIDEBAR
═══════════════════════════════════════════════════════════ */
section[data-testid="stSidebar"] {
  background: var(--bg-surface) !important;
  border-right: 1px solid var(--border-sm) !important;
  box-shadow: 2px 0 12px rgba(10,22,40,0.05) !important;
}

/* Sidebar filter labels */
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .stSelectbox label,
section[data-testid="stSidebar"] .stDateInput label {
  color: var(--tx-3) !important;
  font-size: 10px !important;
  font-weight: 600 !important;
  letter-spacing: 0.08em !important;
  text-transform: uppercase !important;
}

/* Sidebar selectbox */
section[data-testid="stSidebar"] .stSelectbox > div > div {
  background: var(--bg-inset) !important;
  border: 1px solid var(--border-sm) !important;
  border-radius: var(--r-md) !important;
  color: var(--tx-1) !important;
  font-size: 13px !important;
  font-weight: 500 !important;
  box-shadow: var(--shadow-inset) !important;
  transition: var(--t-fast) !important;
}
section[data-testid="stSidebar"] .stSelectbox > div > div:hover {
  border-color: var(--border-md) !important;
  background: var(--bg-surface) !important;
}

/* Sidebar date inputs */
[data-testid="stSidebar"] [data-baseweb="input"] {
  background: var(--bg-inset) !important;
  border-color: var(--border-sm) !important;
  border-radius: var(--r-md) !important;
  box-shadow: var(--shadow-inset) !important;
  font-size: 12px !important;
}

/* Sidebar section headers */
.sidebar-section {
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  color: var(--tx-4);
  margin: 18px 0 8px 2px;
  padding-bottom: 4px;
  border-bottom: 1px solid var(--border-xs);
}

/* ═══════════════════════════════════════════════════════════
   HEADER BAND
═══════════════════════════════════════════════════════════ */
.header-band {
  background: linear-gradient(135deg, var(--navy-900) 0%, var(--navy-700) 55%, var(--navy-600) 100%);
  padding: 16px 28px 16px 28px;
  border-bottom: 1px solid rgba(255,255,255,0.07);
  box-shadow: 0 4px 20px rgba(8,21,40,0.4);
}

.header-top {
  display: flex;
  align-items: flex-start;
  justify-content: space-between;
  margin-bottom: 14px;
  min-height: 44px;
}

.header-logo-area { display: flex; align-items: center; gap: 12px; }

.header-logo-box {
  width: 38px;
  height: 38px;
  background: rgba(255,255,255,0.12);
  border: 1px solid rgba(255,255,255,0.18);
  border-radius: var(--r-md);
  display: flex;
  align-items: center;
  justify-content: center;
  font-size: 13px;
  font-weight: 800;
  color: #FFFFFF;
  letter-spacing: 0.5px;
  flex-shrink: 0;
}

.header-title {
  font-size: 14px;
  font-weight: 700;
  color: #FFFFFF;
  margin-bottom: 2px;
  letter-spacing: 0.1px;
}

.header-subtitle {
  font-size: 11px;
  color: rgba(255,255,255,0.4);
  font-weight: 400;
  letter-spacing: 0.02em;
}

.header-pills {
  display: flex;
  gap: 6px;
  align-items: center;
  flex-shrink: 0;
  margin-top: 2px;
}

.pill {
  display: inline-flex;
  align-items: center;
  gap: 5px;
  background: rgba(255,255,255,0.08);
  border: 1px solid rgba(255,255,255,0.12);
  border-radius: 20px;
  padding: 3px 10px;
  font-size: 10px;
  font-weight: 500;
  color: rgba(255,255,255,0.6);
  letter-spacing: 0.02em;
}

.pill-dot { width:5px; height:5px; border-radius:50%; background:#4ADE80; display:inline-block; }
.pill-live { background: rgba(74,222,128,0.12); border-color: rgba(74,222,128,0.28); color: #4ADE80; }
.pill-txn  { background: rgba(147,197,253,0.10); border-color: rgba(147,197,253,0.22); color: rgba(147,197,253,0.9); }
.pill-period { background: rgba(255,255,255,0.07); border-color: rgba(255,255,255,0.12); }

/* KPI strip */
.kpi-strip {
  display: flex;
  gap: 0;
  align-items: flex-start;
  flex-wrap: wrap;
}

.kpi-strip-item {
  padding: 0 32px 0 0;
  margin-right: 0;
  border-right: 1px solid rgba(255,255,255,0.10);
}
.kpi-strip-item:last-child { border-right: none; padding-right: 0; }

.kpi-strip-hero {
  padding-right: 32px;
  margin-right: 0;
  border-right: 1px solid rgba(255,255,255,0.18) !important;
  padding-bottom: 2px;
}

.kpi-strip-label {
  font-size: 8.5px;
  font-weight: 700;
  color: rgba(255,255,255,0.35);
  letter-spacing: 0.16em;
  text-transform: uppercase;
  margin-bottom: 4px;
}

.kpi-strip-value {
  font-size: 21px;
  font-weight: 800;
  color: #FFFFFF;
  letter-spacing: -0.6px;
  line-height: 1.1;
}

.kpi-strip-value.pos { color: #4ADE80; }
.kpi-strip-value.neg { color: #F87171; }
.kpi-strip-value.sky { color: #93C5FD; }

.kpi-strip-sub { font-size: 10px; margin-top: 3px; font-weight: 400; line-height: 1.3; }
.kpi-positive  { color: rgba(74,222,128,0.65); }
.kpi-negative  { color: rgba(248,113,113,0.65); }
.kpi-neutral   { color: rgba(255,255,255,0.28); }

/* ═══════════════════════════════════════════════════════════
   NAV BAR (button-injected)
═══════════════════════════════════════════════════════════ */
/* Preserved .nav-bar / .nav-tab for any static uses */
.nav-bar {
  background: var(--bg-surface);
  border-bottom: 1px solid var(--border-xs);
  padding: 0 8px;
  margin: 0 0 16px 0;
  display: flex; gap: 0;
}
.nav-tab {
  padding: 11px 18px; font-size: 13px; font-weight: 500;
  color: var(--tx-3); border-bottom: 3px solid transparent; white-space: nowrap;
}
.nav-tab.active { color: var(--navy-700); font-weight: 700; border-bottom-color: var(--navy-700); }

/* Active nav button override */
.stButton > button[kind="primary"] {
  background: var(--blue-50) !important;
  color: var(--navy-700) !important;
  border-color: var(--navy-700) !important;
  font-weight: 700 !important;
}

/* ═══════════════════════════════════════════════════════════
   CARDS & SURFACES  (Phase 3)
═══════════════════════════════════════════════════════════ */
.card {
  background: var(--bg-surface);
  border-radius: var(--r-lg);
  padding: 18px 22px;
  border: 1px solid var(--border-sm);
  box-shadow: var(--shadow-sm);
  margin-bottom: 12px;
  transition: var(--t-base);
  overflow: hidden;
  position: relative;
}
.card:hover {
  box-shadow: var(--shadow-md);
  border-color: var(--border-md);
}

.card-title {
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  color: var(--tx-4);
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  gap: 6px;
}
.card-title::before {
  content: '';
  display: inline-block;
  width: 3px;
  height: 11px;
  background: var(--blue-400);
  border-radius: 2px;
  flex-shrink: 0;
}

/* Chart card wrapper — use around plotly charts */
.chart-card {
  background: var(--bg-surface);
  border-radius: var(--r-lg);
  padding: 16px 18px 12px;
  border: 1px solid var(--border-sm);
  box-shadow: var(--shadow-sm);
  margin-bottom: 12px;
  overflow: hidden;
}

/* ═══════════════════════════════════════════════════════════
   PLOTLY CHART CONTAINERS  (Phase 3)
═══════════════════════════════════════════════════════════ */
[data-testid="stPlotlyChart"] {
  background: var(--bg-surface) !important;
  border-radius: var(--r-lg) !important;
  border: 1px solid var(--border-sm) !important;
  box-shadow: var(--shadow-sm) !important;
  padding: 6px 4px 2px !important;
  overflow: hidden !important;
}

/* ═══════════════════════════════════════════════════════════
   ENTITY CARDS  (Phase 3)
═══════════════════════════════════════════════════════════ */
.entity-card {
  background: var(--bg-surface);
  border-radius: var(--r-lg);
  padding: 0 0 16px 0;
  border: 1px solid var(--border-sm);
  box-shadow: var(--shadow-sm);
  margin-bottom: 10px;
  transition: var(--t-base);
  overflow: hidden;
}
.entity-card:hover {
  box-shadow: var(--shadow-md);
  border-color: var(--border-md);
  transform: translateY(-1px);
}

/* Colored top accent bar — first entity blue, second teal */
.entity-card:nth-child(1) { border-top: 3px solid var(--blue-400); }
.entity-card:nth-child(2) { border-top: 3px solid #14B8A6; }

.entity-name {
  font-size: 9.5px;
  font-weight: 700;
  color: var(--tx-2);
  letter-spacing: 0.13em;
  text-transform: uppercase;
  padding: 12px 18px 10px;
  border-bottom: 1px solid var(--border-xs);
  background: var(--bg-raised);
}

.stat-row { display: flex; gap: 10px; flex-wrap: wrap; padding: 12px 16px 0; }

.stat-badge {
  display: inline-flex;
  flex-direction: column;
  background: var(--bg-raised);
  border-radius: var(--r-md);
  padding: 8px 14px;
  min-width: 88px;
  border: 1px solid var(--border-xs);
  transition: var(--t-fast);
}
.stat-badge:hover {
  border-color: var(--border-md);
  background: var(--bg-surface);
  box-shadow: var(--shadow-xs);
}

.stat-badge-label {
  font-size: 8px;
  font-weight: 700;
  letter-spacing: 0.13em;
  text-transform: uppercase;
  color: var(--tx-4);
  margin-bottom: 4px;
}

.stat-badge-value { font-size: 14px; font-weight: 700; color: var(--tx-1); letter-spacing: -0.3px; }
.badge-green .stat-badge-value { color: var(--green-700); }
.badge-red   .stat-badge-value { color: var(--red-700); }
.badge-blue  .stat-badge-value { color: var(--blue-600); }
.badge-grey  .stat-badge-value { color: var(--tx-3); }

/* ═══════════════════════════════════════════════════════════
   SECTION HEADERS
═══════════════════════════════════════════════════════════ */
.section-title, .section-header {
  font-size: 10px;
  font-weight: 700;
  color: var(--tx-4);
  letter-spacing: 0.12em;
  text-transform: uppercase;
  margin: 20px 0 10px 0;
}

/* ═══════════════════════════════════════════════════════════
   KPI CARDS  (Phase 3 — refined)
═══════════════════════════════════════════════════════════ */
.kpi-card {
  background: var(--bg-surface);
  border-radius: var(--r-lg);
  padding: 16px 20px 16px 22px;
  border: 1px solid var(--border-sm);
  box-shadow: var(--shadow-sm);
  position: relative;
  overflow: hidden;
  margin-bottom: 10px;
  transition: var(--t-base);
}
.kpi-card:hover {
  box-shadow: var(--shadow-md);
  border-color: var(--border-md);
}
.kpi-card::before {
  content: ''; position: absolute; top: 0; left: 0;
  width: 4px; height: 100%; border-radius: var(--r-lg) 0 0 var(--r-lg);
}
.kpi-card.green::before { background: linear-gradient(180deg, var(--green-400), var(--green-700)); }
.kpi-card.red::before   { background: linear-gradient(180deg, var(--red-400),   var(--red-700)); }
.kpi-card.blue::before  { background: linear-gradient(180deg, #60A5FA,          var(--blue-600)); }
.kpi-card.navy::before  { background: linear-gradient(180deg, var(--navy-500),  var(--navy-900)); }

.kpi-label {
  font-size: 9px; font-weight: 700; color: var(--tx-4);
  letter-spacing: 0.13em; text-transform: uppercase; margin-bottom: 8px;
}
.kpi-value {
  font-size: 24px; font-weight: 800; color: var(--tx-1);
  margin-bottom: 3px; letter-spacing: -0.6px; line-height: 1.1;
}
.kpi-value.positive { color: var(--green-700); }
.kpi-value.negative { color: var(--red-700); }
.kpi-sub { font-size: 11px; color: var(--tx-4); margin-top: 4px; line-height: 1.4; }

/* ═══════════════════════════════════════════════════════════
   TABLES & DATA FRAMES  (Phase 3)
═══════════════════════════════════════════════════════════ */
.stDataFrame,
[data-testid="stDataFrameResizable"] {
  border: 1px solid var(--border-sm) !important;
  border-radius: var(--r-lg) !important;
  overflow: hidden !important;
  box-shadow: var(--shadow-sm) !important;
  background: var(--bg-surface) !important;
}

/* Header row */
.stDataFrame thead tr th {
  background: linear-gradient(180deg, var(--bg-raised) 0%, #ECEEF3 100%) !important;
  color: var(--tx-3) !important;
  font-size: 9.5px !important;
  font-weight: 700 !important;
  letter-spacing: 0.09em !important;
  text-transform: uppercase !important;
  border-bottom: 2px solid var(--border-sm) !important;
  padding: 7px 12px !important;
  white-space: nowrap !important;
}

/* Body rows */
.stDataFrame tbody tr td {
  font-size: 12.5px !important;
  color: var(--tx-1) !important;
  padding: 6px 12px !important;
  border-bottom: 1px solid var(--border-xs) !important;
  transition: background 0.1s ease !important;
}
.stDataFrame tbody tr:nth-child(even) td {
  background: var(--bg-raised) !important;
}
.stDataFrame tbody tr:hover td {
  background: var(--blue-50) !important;
  color: var(--navy-700) !important;
}

/* ═══════════════════════════════════════════════════════════
   HR DIVIDERS
═══════════════════════════════════════════════════════════ */
hr { border-color: var(--border-xs) !important; margin: 10px 0 !important; }

/* ═══════════════════════════════════════════════════════════
   BUTTONS
═══════════════════════════════════════════════════════════ */
.stButton > button {
  background: var(--bg-surface) !important;
  color: var(--tx-2) !important;
  border: 1px solid var(--border-sm) !important;
  border-radius: var(--r-md) !important;
  font-size: 12.5px !important;
  font-weight: 500 !important;
  padding: 5px 14px !important;
  transition: var(--t-fast) !important;
  box-shadow: var(--shadow-xs) !important;
  letter-spacing: 0.01em !important;
}
.stButton > button:hover {
  background: var(--bg-inset) !important;
  border-color: var(--border-md) !important;
  color: var(--navy-700) !important;
  box-shadow: var(--shadow-sm) !important;
}
.stButton > button:active {
  transform: translateY(1px) !important;
  box-shadow: none !important;
}

[data-testid="stDownloadButton"] > button {
  background: var(--navy-700) !important;
  color: #FFFFFF !important;
  border-color: var(--navy-700) !important;
  border-radius: var(--r-md) !important;
  font-weight: 600 !important;
  box-shadow: 0 1px 4px rgba(27,43,75,0.3) !important;
}
[data-testid="stDownloadButton"] > button:hover {
  background: var(--navy-800) !important;
  box-shadow: 0 2px 8px rgba(27,43,75,0.4) !important;
}

/* ═══════════════════════════════════════════════════════════
   FORM INPUTS
═══════════════════════════════════════════════════════════ */
[data-testid="stTextInput"] > div > div {
  border-radius: var(--r-md) !important;
  border-color: var(--border-sm) !important;
  background: var(--bg-surface) !important;
  font-size: 13px !important;
  box-shadow: var(--shadow-inset) !important;
}
[data-testid="stTextInput"] > div > div:focus-within {
  border-color: var(--blue-400) !important;
  box-shadow: 0 0 0 3px rgba(59,130,246,0.12) !important;
}

[data-testid="stSelectbox"] > div > div {
  border-radius: var(--r-md) !important;
  border-color: var(--border-sm) !important;
  font-size: 13px !important;
}

[data-baseweb="select"] {
  border-radius: var(--r-md) !important;
}

/* ═══════════════════════════════════════════════════════════
   EXPANDERS
═══════════════════════════════════════════════════════════ */
[data-testid="stExpander"] {
  background: var(--bg-surface) !important;
  border-radius: var(--r-md) !important;
  border: 1px solid var(--border-sm) !important;
  box-shadow: var(--shadow-xs) !important;
  margin-bottom: 8px !important;
}
[data-testid="stExpander"] summary {
  font-size: 12.5px !important;
  font-weight: 600 !important;
  color: var(--tx-2) !important;
  padding: 10px 14px !important;
}
[data-testid="stExpander"] summary:hover {
  background: var(--bg-raised) !important;
  border-radius: var(--r-md) var(--r-md) 0 0 !important;
}

/* ═══════════════════════════════════════════════════════════
   METRICS
═══════════════════════════════════════════════════════════ */
[data-testid="stMetric"] {
  background: var(--bg-surface) !important;
  border-radius: var(--r-md) !important;
  padding: 12px 16px !important;
  border: 1px solid var(--border-sm) !important;
  box-shadow: var(--shadow-xs) !important;
}
[data-testid="stMetricValue"] { color: var(--tx-1) !important; }
[data-testid="stMetricLabel"] { color: var(--tx-4) !important; font-size: 10px !important; }

/* ═══════════════════════════════════════════════════════════
   CAPTIONS & ALERTS
═══════════════════════════════════════════════════════════ */
[data-testid="stCaptionContainer"] p {
  color: var(--tx-3) !important;
  font-size: 11.5px !important;
}

[data-testid="stAlert"] { border-radius: var(--r-md) !important; }

.alert-box {
  border-radius: var(--r-md);
  padding: 12px 16px;
  margin: 8px 0;
  border-left: 3px solid;
  font-size: 13px;
}
.alert-red    { background: var(--red-50);   border-color: var(--red-500);   color: #7F1D1D; }
.alert-yellow { background: var(--amber-50); border-color: var(--amber-700); color: #78350F; }
.alert-green  { background: var(--green-50); border-color: var(--green-500); color: #14532D; }
.alert-blue   { background: var(--blue-50);  border-color: var(--blue-400);  color: var(--navy-700); }

/* ═══════════════════════════════════════════════════════════
   SIDEBAR STAT CARDS
═══════════════════════════════════════════════════════════ */
.sb-stats-grid {
  display: flex;
  flex-direction: column;
  gap: 7px;
  margin-bottom: 12px;
}

.sb-stat-card {
  background: var(--bg-raised);
  border-radius: var(--r-md);
  padding: 9px 14px;
  border: 1px solid var(--border-xs);
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 8px;
}
.sb-stat-card.green { background: var(--green-50);  border-color: var(--green-border); }
.sb-stat-card.red   { background: var(--red-50);    border-color: var(--red-border); }
.sb-stat-card.amber { background: var(--amber-50);  border-color: var(--amber-border); }

.sb-stat-left { display: flex; flex-direction: column; gap: 2px; }

.sb-stat-label {
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 0.13em;
  text-transform: uppercase;
  color: var(--tx-4);
}
.sb-stat-label.green { color: var(--green-700); }
.sb-stat-label.red   { color: var(--red-700); }
.sb-stat-label.amber { color: var(--amber-700); }

.sb-stat-value {
  font-size: 15px;
  font-weight: 700;
  color: var(--tx-1);
  letter-spacing: -0.3px;
}
.sb-stat-value.green { color: var(--green-700); }
.sb-stat-value.red   { color: var(--red-700); }
.sb-stat-value.amber { color: var(--amber-700); }

.sb-stat-icon {
  font-size: 18px;
  opacity: 0.45;
  flex-shrink: 0;
}

/* ═══════════════════════════════════════════════════════════
   WoW TREND BADGES
═══════════════════════════════════════════════════════════ */
.kpi-wow {
  display: inline-flex;
  align-items: center;
  gap: 3px;
  font-size: 9.5px;
  font-weight: 600;
  padding: 2px 7px;
  border-radius: 20px;
  margin-top: 5px;
  letter-spacing: 0.02em;
}
.kpi-wow.up   { background: rgba(74,222,128,0.18); color: #4ADE80; }
.kpi-wow.down { background: rgba(248,113,113,0.18); color: #F87171; }
.kpi-wow.flat { background: rgba(255,255,255,0.08); color: rgba(255,255,255,0.4); }
.kpi-wow.warn { background: rgba(251,191,36,0.18);  color: #FBBF24; }

/* ═══════════════════════════════════════════════════════════
   EXECUTIVE INSIGHTS PANEL
═══════════════════════════════════════════════════════════ */
.insight-panel {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(240px, 1fr));
  gap: 10px;
  margin: 0 0 20px;
}
.insight-card {
  background: var(--bg-surface);
  border-radius: var(--r-lg);
  border: 1px solid var(--border-sm);
  padding: 13px 15px;
  display: flex;
  gap: 11px;
  align-items: flex-start;
  box-shadow: var(--shadow-xs);
}
.insight-card.green { border-color: var(--green-border); background: var(--green-50); }
.insight-card.red   { border-color: var(--red-border);   background: var(--red-50); }
.insight-card.amber { border-color: var(--amber-border); background: var(--amber-50); }
.insight-card.blue  { border-color: #BFDBFE; background: #EFF6FF; }
.insight-icon { font-size: 17px; line-height: 1.1; flex-shrink: 0; }
.insight-body { display: flex; flex-direction: column; gap: 3px; }
.insight-title { font-size: 11.5px; font-weight: 600; color: var(--tx-1); line-height: 1.35; }
.insight-desc  { font-size: 10.5px; color: var(--tx-3); line-height: 1.45; }

/* ═══════════════════════════════════════════════════════════
   ENTITY BREAKDOWN CARDS
═══════════════════════════════════════════════════════════ */
.ent-cards {
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(210px, 1fr));
  gap: 12px;
  margin: 0 0 18px;
}
.ent-card {
  background: var(--bg-surface);
  border-radius: var(--r-lg);
  border: 1px solid var(--border-sm);
  padding: 16px 20px;
  box-shadow: var(--shadow-xs);
  transition: var(--t-base);
}
.ent-card:hover { box-shadow: var(--shadow-md); border-color: var(--border-md); }
.ent-card-name { font-size: 9px; font-weight: 700; letter-spacing: 0.13em; text-transform: uppercase; color: var(--tx-4); margin-bottom: 7px; }
.ent-card-net  { font-size: 27px; font-weight: 800; letter-spacing: -0.6px; line-height: 1; margin-bottom: 10px; }
.ent-card-net.pos { color: var(--green-700); }
.ent-card-net.neg { color: var(--red-700); }
.ent-card-row  { display: flex; justify-content: space-between; align-items: baseline; font-size: 11px; color: var(--tx-3); padding: 3px 0; border-top: 1px solid var(--border-xs); }
.ent-card-in   { color: var(--green-700); font-weight: 600; font-size: 12px; }
.ent-card-out  { color: var(--red-700);   font-weight: 600; font-size: 12px; }

/* ═══════════════════════════════════════════════════════════
   KPI STRIP ENHANCEMENTS
═══════════════════════════════════════════════════════════ */
.kpi-strip-divider {
  height: 1px;
  background: rgba(255,255,255,0.08);
  margin: 12px 0 14px;
}

.kpi-arrow-up   { font-size: 10px; color: rgba(74,222,128,0.8); margin-right: 2px; }
.kpi-arrow-down { font-size: 10px; color: rgba(248,113,113,0.8); margin-right: 2px; }

/* ═══════════════════════════════════════════════════════════
   CASH FLOW STATEMENT  (Phase 4)
═══════════════════════════════════════════════════════════ */

/* Control bar (filter strip) */
.cf-control-bar {
  background: linear-gradient(135deg, var(--navy-900) 0%, var(--navy-700) 60%, var(--navy-600) 100%);
  border-radius: var(--r-lg);
  padding: 16px 22px 10px;
  margin-bottom: 10px;
  border: 1px solid rgba(255,255,255,0.06);
  box-shadow: var(--shadow-md);
}

.cf-bar-label {
  display: flex;
  align-items: center;
  gap: 10px;
  color: rgba(255,255,255,0.7);
  font-size: 13px;
  font-weight: 700;
  letter-spacing: 0.08em;
  text-transform: uppercase;
  margin-bottom: 14px;
}
.cf-bar-label::after {
  content: '';
  flex: 1;
  height: 1px;
  background: rgba(255,255,255,0.1);
}

/* Section dividers */
.cf-divider-sm { border: none !important; border-top: 1px solid var(--border-xs) !important; margin: 2px 0 !important; }
.cf-divider-lg { border: none !important; border-top: 2px solid var(--blue-400)  !important; margin: 2px 0 !important; }

/* Section header labels */
.cf-section-label {
  font-weight: 700;
  font-size: 14px;
  padding: 6px 10px;
  border-radius: var(--r-sm);
  display: block;
  line-height: 1.3;
}
.cf-ob  { background: #DBEAFE; color: #1E40AF; border-left: 3px solid #3B82F6; }
.cf-rec { background: #EDE9FE; color: #5B21B6; border-left: 3px solid #7C3AED; }
.cf-pay { background: #FCE7F3; color: #9D174D; border-left: 3px solid #EC4899; }
.cf-cb  { background: #DBEAFE; color: #1E40AF; border-left: 3px solid #3B82F6; }

/* Detail rows */
.cf-detail     { padding: 4px 8px 4px 28px; font-size: 13px; color: var(--tx-2); }
.cf-detail.unc { font-style: italic; color: var(--tx-3); }
.cf-dval       { text-align: right; font-size: 13px; color: var(--tx-2); padding: 4px 0; }
.cf-dval.unc   { color: var(--tx-3); }

/* NET CASH FLOW hero banner */
.cf-net {
  display: flex;
  justify-content: space-between;
  align-items: center;
  font-weight: 800;
  font-size: 16px;
  letter-spacing: 0.01em;
  padding: 13px 16px;
  margin-top: 10px;
  border-radius: var(--r-md);
  border-top: 3px solid;
  border-bottom: 3px solid;
  box-shadow: var(--shadow-xs);
}
.cf-net.pos { border-color: var(--green-500); color: var(--green-700); background: var(--green-50); }
.cf-net.neg { border-color: var(--red-500);   color: var(--red-700);   background: var(--red-50); }
.cf-net span:last-child { font-size: 18px; letter-spacing: -0.5px; }

/* Tally check pill */
.cf-tally {
  border-radius: var(--r-md);
  padding: 7px 14px;
  font-size: 11.5px;
  margin-top: 6px;
  display: flex;
  align-items: center;
  gap: 8px;
  font-weight: 500;
  line-height: 1.5;
  flex-wrap: wrap;
}
.cf-tally.ok  { background: var(--green-50); color: var(--green-700); border: 1px solid var(--green-border); }
.cf-tally.err { background: var(--red-50);   color: var(--red-700);   border: 1px solid var(--red-border); }

/* ── Statement context band ─────────────────────────────────────────────────── */
.cf-stmt-band {
  display: flex;
  justify-content: space-between;
  align-items: center;
  background: var(--bg-raised);
  border: 1px solid var(--border-sm);
  border-radius: var(--r-md);
  padding: 8px 14px;
  margin-bottom: 10px;
  font-size: 11px;
  color: var(--tx-3);
  gap: 16px;
}
.cf-stmt-band-kpi {
  display: flex;
  align-items: baseline;
  gap: 6px;
}
.cf-stmt-band-label { font-size: 9px; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase; color: var(--tx-4); }
.cf-stmt-band-val   { font-size: 13px; font-weight: 700; color: var(--tx-1); letter-spacing: -0.3px; }
.cf-stmt-band-val.pos { color: var(--green-700); }
.cf-stmt-band-val.neg { color: var(--red-700); }

/* ── Section total value cells ─────────────────────────────────────────────── */
.cf-section-total {
  font-weight: 700;
  text-align: right;
  padding: 6px 0;
  font-size: 15px;
  letter-spacing: -0.4px;
  color: var(--tx-1);
  line-height: 1.2;
}
.cf-section-total.pos { color: var(--green-700); }
.cf-section-total.neg { color: var(--red-700); }
.cf-section-total.sky { color: var(--navy-700); }

/* ── GROUP level rows (L2) ──────────────────────────────────────────────────── */
.cf-grp-row {
  font-weight: 600;
  font-size: 12.5px;
  padding: 5px 10px 5px 10px;
  border-radius: 5px;
  margin: 1px 0;
  letter-spacing: 0.005em;
  display: block;
}
.cf-grp-row.rec { background: #EEF2FF; color: #1E40AF; }
.cf-grp-row.pay { background: #FFF1F2; color: #9D174D; }

/* GROUP total value cells */
.cf-grp-val {
  font-weight: 600;
  font-size: 12.5px;
  text-align: right;
  padding: 5px 2px;
}
.cf-grp-val.rec { color: #1E40AF; }
.cf-grp-val.pay { color: #9D174D; }

/* ── FINAL GROUP sub-row label (L3) ─────────────────────────────────────────── */
.cf-sub-row {
  font-size: 12px;
  color: var(--tx-2);
  padding: 3px 8px 3px 32px;
  border-radius: 3px;
  display: block;
}
.cf-sub-row.unc { font-style: italic; color: var(--tx-3); }

/* FINAL GROUP sub-row value */
.cf-sub-val {
  font-size: 12px;
  color: var(--tx-2);
  text-align: right;
  padding: 3px 2px;
}
.cf-sub-val.unc { color: var(--tx-3); }

/* ── Weekly GROUP label cells ─────────────────────────────────────────────── */
.cf-wk-grp-lbl {
  font-weight: 600;
  font-size: 12px;
  padding: 4px 8px;
  border-radius: 4px;
  display: block;
}
.cf-wk-grp-lbl.rec { background: #EEF2FF; color: #1E40AF; }
.cf-wk-grp-lbl.pay { background: #FFF1F2; color: #9D174D; }

/* Weekly GROUP value cells */
.cf-wk-grp-val {
  text-align: right;
  font-size: 12px;
  font-weight: 600;
  padding: 4px 3px;
}
.cf-wk-grp-val.rec       { color: #1E40AF; }
.cf-wk-grp-val.rec.total { background: #EFF6FF; color: #1E40AF; }
.cf-wk-grp-val.pay       { color: #9D174D; }
.cf-wk-grp-val.pay.total { background: #FFF1F2; color: #9D174D; }

/* ── Weekly sub-row cells ─────────────────────────────────────────────────── */
.cf-wk-sub-lbl {
  font-size: 12px;
  color: var(--tx-2);
  padding: 3px 8px 3px 24px;
  display: block;
}
.cf-wk-sub-lbl.unc { font-style: italic; color: var(--tx-3); }

.cf-wk-sub-val {
  text-align: right;
  font-size: 12px;
  color: var(--tx-2);
  padding: 3px 3px;
}
.cf-wk-sub-val.unc   { color: var(--tx-3); }
.cf-wk-sub-val.total { font-weight: 600; color: #1E40AF; background: #EFF6FF; }
.cf-wk-sub-val.ptotal { font-weight: 600; color: #9D174D; background: #FFF1F2; }

/* Weekly column header cells */
.cf-wk-hdr {
  font-size: 13px;
  font-weight: 700;
  text-align: right;
  padding: 6px 4px;
  border-radius: var(--r-sm);
  letter-spacing: 0.02em;
}
.cf-wk-hdr.alt { background: #DBEAFE; color: var(--blue-600); }  /* kept for possible future use */
.cf-wk-hdr.std { background: var(--bg-raised); color: var(--tx-2); }
.cf-wk-hdr.ttl { background: var(--blue-50); color: var(--navy-700); }

/* Weekly cell values */
.cf-wk-val {
  text-align: right;
  font-size: 13px;
  font-weight: 500;
  color: var(--tx-1);
  padding: 4px 4px;
}
.cf-wk-val.bold   { font-weight: 700; }
.cf-wk-val.red    { color: var(--red-700, #B91C1C); }
.cf-wk-val.total  { background: var(--blue-50); font-weight: 600; }
.cf-wk-val.dim    { color: var(--tx-3); }  /* for — dashes */

/* Weekly NET label cell */
.cf-wk-net-label {
  font-weight: 800;
  font-size: 13.5px;
  padding: 7px 10px;
  border-radius: var(--r-sm);
  background: var(--green-50);
  border-left: 3px solid var(--green-500);
  color: var(--green-700);
  display: block;
}

/* Weekly tally label */
.cf-wk-tally-label {
  color: var(--tx-3);
  font-size: 12px;
  padding: 4px 8px;
}

/* ═══════════════════════════════════════════════════════════
   TRANSACTIONS TAB
═══════════════════════════════════════════════════════════ */

/* Filter bar — same dark-navy pattern as CF control bar */
.txn-filter-bar {
  background: linear-gradient(135deg, var(--navy-900) 0%, var(--navy-700) 60%, var(--navy-600) 100%);
  border-radius: var(--r-lg);
  padding: 14px 20px 10px;
  margin-bottom: 12px;
  border: 1px solid rgba(255,255,255,0.06);
  box-shadow: var(--shadow-md);
}
.txn-filter-title {
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 0.16em;
  text-transform: uppercase;
  color: rgba(255,255,255,0.45);
  margin-bottom: 10px;
  display: flex;
  align-items: center;
  gap: 10px;
}
.txn-filter-title::after {
  content: '';
  flex: 1;
  height: 1px;
  background: rgba(255,255,255,0.08);
}

/* Transaction count badge */
.txn-count-badge {
  display: inline-flex;
  align-items: center;
  gap: 7px;
  background: var(--navy-700);
  color: #FFFFFF;
  font-size: 12px;
  font-weight: 600;
  padding: 5px 14px;
  border-radius: 20px;
  margin: 6px 0 10px;
  letter-spacing: 0.02em;
}
.txn-count-badge .txn-dot {
  width: 6px;
  height: 6px;
  border-radius: 50%;
  background: #4ADE80;
  flex-shrink: 0;
}

/* Transaction intelligence strip */
.txn-intel-strip {
  display: flex;
  background: var(--bg-surface);
  border: 1px solid var(--border-sm);
  border-radius: var(--r-md);
  overflow: hidden;
  margin-bottom: 14px;
  box-shadow: var(--shadow-xs);
}
.txn-intel-item {
  flex: 1;
  padding: 10px 14px;
  border-right: 1px solid var(--border-xs);
}
.txn-intel-item:last-child { border-right: none; }
.txn-intel-label {
  font-size: 8.5px;
  font-weight: 700;
  letter-spacing: 0.13em;
  text-transform: uppercase;
  color: var(--tx-4);
  margin-bottom: 3px;
}
.txn-intel-val {
  font-size: 15px;
  font-weight: 700;
  color: var(--tx-1);
  letter-spacing: -0.4px;
  line-height: 1.1;
}
.txn-intel-val.pos   { color: var(--green-700); }
.txn-intel-val.neg   { color: var(--red-700); }
.txn-intel-val.amber { color: var(--amber-700); }
.txn-intel-sub {
  font-size: 10px;
  color: var(--tx-4);
  margin-top: 2px;
  line-height: 1.3;
}

/* Bulk categorization panel */
.txn-cat-panel {
  background: var(--bg-raised);
  border: 1px solid var(--border-sm);
  border-radius: var(--r-md);
  padding: 12px 16px 8px;
  margin: 4px 0 10px;
}
.txn-cat-title {
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 0.13em;
  text-transform: uppercase;
  color: var(--tx-4);
  margin-bottom: 8px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.txn-cat-title::after {
  content: '';
  flex: 1;
  height: 1px;
  background: var(--border-xs);
}

/* Opening Balance row banner — enhanced */
.ob-banner {
  background: var(--blue-50);
  border: 1px solid #BFDBFE;
  border-radius: var(--r-md);
  padding: 8px 16px;
  margin-bottom: 8px;
  display: flex;
  align-items: center;
  gap: 8px;
}

/* ═══════════════════════════════════════════════════════════
   POLISH COMPONENTS  (Phase 5)
═══════════════════════════════════════════════════════════ */

/* Streamlit h3 / h4 override — keep consistent with design */
.main h1, .main h2, .main h3, .main h4, .main h5 {
  font-family: 'Inter', sans-serif !important;
  font-weight: 700 !important;
  letter-spacing: -0.2px !important;
}
.main h3 { font-size: 15px !important; color: var(--tx-1) !important; margin: 18px 0 10px !important; }
.main h4 {
  font-size: 11px !important;
  color: var(--tx-3) !important;
  text-transform: uppercase !important;
  letter-spacing: 0.11em !important;
  margin: 16px 0 8px !important;
  padding-bottom: 7px !important;
  border-bottom: 1px solid var(--border-xs) !important;
}

/* Opening Balance row banner (Transactions tab) */
.ob-banner {
  background: var(--blue-50);
  border-radius: var(--r-md);
  padding: 8px 14px;
  border-left: 4px solid var(--blue-400);
  margin-bottom: 6px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.ob-banner-label {
  font-size: 9.5px;
  font-weight: 700;
  letter-spacing: 0.13em;
  text-transform: uppercase;
  color: var(--blue-500);
}

/* Legend badges */
.legend-strip {
  display: flex;
  gap: 10px;
  flex-wrap: wrap;
  margin: 8px 0;
}
.legend-badge {
  padding: 4px 11px;
  border-radius: var(--r-sm);
  border: 1px solid;
  font-size: 12px;
  font-weight: 500;
  display: inline-flex;
  align-items: center;
  gap: 5px;
}
.legend-badge.green { background: var(--green-50);  border-color: var(--green-border);  color: var(--green-700); }
.legend-badge.amber { background: var(--amber-50);  border-color: var(--amber-border);  color: var(--amber-700); }

/* Reconciliation strip */
.recon-strip {
  background: var(--bg-raised);
  border-radius: var(--r-md);
  padding: 11px 18px;
  border: 1px solid var(--border-sm);
  margin-top: 10px;
  display: flex;
  flex-wrap: wrap;
  gap: 4px 18px;
  align-items: center;
}
.recon-label {
  font-size: 9px;
  font-weight: 700;
  letter-spacing: 0.13em;
  text-transform: uppercase;
  color: var(--tx-4);
  width: 100%;
  margin-bottom: 3px;
}
.recon-item       { font-size: 13px; color: var(--tx-2); }
.recon-item b     { font-weight: 700; color: var(--tx-1); }
.recon-item b.pos { color: var(--green-700); }
.recon-item b.neg { color: var(--red-700); }
.recon-sep        { color: var(--border-md); font-size: 14px; }

/* ═══════════════════════════════════════════════════════════
   STREAMLIT CHROME CLEANUP  (Phase 5 refined)
═══════════════════════════════════════════════════════════ */
/* Tighten main area top padding */
.main .block-container { padding-top: 0 !important; }

/* Remove extra whitespace above first element */
.main > div:first-child { padding-top: 0 !important; margin-top: 0 !important; }

/* Tighten column gaps */
[data-testid="stHorizontalBlock"] { gap: 12px !important; }

/* Caption below inputs */
.stCaption { margin-top: 2px !important; }

/* Streamlit info / success / warning / error boxes */
[data-testid="stInfoBox"], [data-testid="stSuccessBox"],
[data-testid="stWarningBox"], [data-testid="stErrorBox"] {
  border-radius: var(--r-md) !important;
  font-size: 13px !important;
  font-weight: 500 !important;
}

/* File uploader */
[data-testid="stFileUploader"] {
  border-radius: var(--r-lg) !important;
  border: 1.5px dashed var(--border-md) !important;
  background: var(--bg-raised) !important;
  padding: 6px !important;
  transition: var(--t-fast) !important;
}
[data-testid="stFileUploader"]:hover {
  border-color: var(--blue-400) !important;
  background: var(--blue-50) !important;
}

/* Radio buttons */
[data-testid="stRadio"] label {
  font-size: 13px !important;
  font-weight: 500 !important;
}

/* Date inputs */
[data-testid="stDateInput"] input {
  font-size: 13px !important;
}

/* Checkbox */
[data-testid="stCheckbox"] label {
  font-size: 13px !important;
}

/* Tooltip / popover */
[data-testid="stTooltipIcon"] { color: var(--tx-4) !important; }

/* ═══════════════════════════════════════════════════════════
   RESPONSIVE  (Phase 5)
═══════════════════════════════════════════════════════════ */
@media (max-width: 1200px) {
  .header-band { padding: 14px 20px 14px; }
  .kpi-strip-item { padding-right: 22px; }
  .kpi-strip-value { font-size: 18px; }
  .header-pills .pill-period { display: none; }
}

@media (max-width: 960px) {
  .kpi-strip { flex-wrap: wrap; row-gap: 14px; }
  .kpi-strip-item {
    padding-right: 16px;
    border-right: none !important;
    min-width: calc(50% - 16px);
  }
  .kpi-strip-value { font-size: 17px; }
  .header-pills { display: none; }
  .header-subtitle { font-size: 10px; }
}

@media (max-width: 700px) {
  .header-top { flex-direction: column; gap: 10px; }
  .kpi-strip-item { min-width: 100%; padding-right: 0; }
  .kpi-strip-value { font-size: 19px; }
  .header-logo-box { width: 32px; height: 32px; font-size: 11px; }
  .header-title { font-size: 13px; }
  .recon-strip { flex-direction: column; gap: 6px; }
  .legend-strip { flex-direction: column; }
  .sb-stats-grid { flex-direction: row; flex-wrap: wrap; }
  .sb-stat-card { min-width: calc(50% - 4px); }
}

/* ═══════════════════════════════════════════════════════════
   SIDEBAR LOGO GAP REDUCTION
═══════════════════════════════════════════════════════════ */
section[data-testid="stSidebar"] > div {
  padding-top: 0.25rem !important;
}
section[data-testid="stSidebar"] [data-testid="stImage"] {
  margin-top: 0 !important;
  margin-bottom: 0 !important;
  padding-top: 0 !important;
  padding-bottom: 0 !important;
  line-height: 0 !important;
}
section[data-testid="stSidebar"] [data-testid="stImage"] img {
  display: block !important;
  margin: 0 auto !important;
}
</style>
""", unsafe_allow_html=True)

init_db()

# ── Navigation session state ──────────────────────────────────────────────────
_NAV_TABS = ["Summary", "Cash Flow", "Transactions", "Uncategorised", "Exception Report", "Investments", "Upload"]

_today          = datetime.date.today()
_first_of_month = _today.replace(day=1)

# ─── Sidebar ───────────────────────────────────────────────────────────────────
import os as _os, base64 as _b64
LOGO_PATH = _os.path.join(_os.path.dirname(__file__), "logo.png")
_logo_display_path = LOGO_PATH
if not _os.path.exists(LOGO_PATH):
    _fallback = _os.path.join(_os.path.dirname(__file__), "citykart_logo.png")
    _logo_display_path = _fallback if _os.path.exists(_fallback) else None

if _logo_display_path:
    with open(_logo_display_path, "rb") as _lf:
        _logo_b64 = _b64.b64encode(_lf.read()).decode()
    st.sidebar.markdown(f"""
<div style="padding:10px 0 14px 0;text-align:center;">
  <img src="data:image/png;base64,{_logo_b64}"
       style="max-width:80%;max-height:56px;object-fit:contain;border-radius:6px;" />
</div>
""", unsafe_allow_html=True)
else:
    st.sidebar.markdown("""
<div style="
    width:100%;height:56px;
    background:var(--bg-inset,#F0F3F8);
    border:1.5px dashed var(--border-md,#C8D3E0);
    border-radius:10px;
    display:flex;align-items:center;justify-content:center;
    gap:8px;margin-bottom:14px;
    color:var(--tx-4,#96A3B4);font-size:11px;font-weight:600;
    letter-spacing:0.08em;text-transform:uppercase;
">
  <span style="opacity:0.5;">🏦</span> Add logo.png
</div>
""", unsafe_allow_html=True)

with st.sidebar.expander("🖼 Update Logo", expanded=False):
    _logo_upload_sb = st.file_uploader(
        "Upload PNG logo", type=["png", "jpg", "jpeg"], key="logo_upload_sb"
    )
    if _logo_upload_sb is not None:
        if st.button("💾 Save Logo", key="logo_save_sb"):
            with open(LOGO_PATH, "wb") as _lw:
                _lw.write(_logo_upload_sb.getvalue())
            st.success("Logo saved as logo.png — persists after restart.")
            st.cache_data.clear()
            try:
                st.rerun()
            except AttributeError:
                st.experimental_rerun()

# ── Derive default FY from latest transaction date in DB ──────────────────────
@st.cache_data(ttl=300)
def _get_latest_date():
    """Return the most recent transaction date (YYYY-MM-DD) in the DB, or None."""
    from database import get_connection
    conn = get_connection()
    row  = conn.execute("""
        SELECT MAX(date) as latest
        FROM transactions
        WHERE final_group != 'OPENING BALANCE'
    """).fetchone()
    conn.close()
    return row["latest"] if row and row["latest"] else None

_latest_date = _get_latest_date()
if _latest_date:
    _ld = datetime.date.fromisoformat(_latest_date)
    if _ld >= datetime.date(2026, 4, 1):
        _default_fy = "FY2627"
    elif _ld >= datetime.date(2025, 4, 1):
        _default_fy = "FY2526"
    else:
        _default_fy = "FY2425"
else:
    _default_fy = "FY2627"
_default_fy_idx = FY_OPTIONS.index(_default_fy)

st.sidebar.markdown('<div class="sidebar-section">Filters</div>', unsafe_allow_html=True)

sel_fy       = st.sidebar.selectbox(
    "Financial Year", FY_OPTIONS,
    index=_default_fy_idx,
    format_func=lambda x: FY_LABELS[x],
)
sel_fy_label = FY_LABELS[sel_fy]

# Reset Cash Flow date inputs whenever the FY selector changes
_fy_s, _fy_e = FY_BOUNDS[sel_fy]
if st.session_state.get("_prev_fy") != sel_fy:
    st.session_state["_prev_fy"] = sel_fy
    st.cache_data.clear()

    # Default CF window to the latest month that has data for this FY
    from database import get_connection as _gc
    _conn = _gc()
    _row  = _conn.execute("""
        SELECT MAX(date) as latest, MIN(date) as earliest
        FROM transactions
        WHERE financial_year = ?
          AND final_group != 'OPENING BALANCE'
    """, [sel_fy]).fetchone()
    _conn.close()

    if _row and _row["latest"]:
        _cf_latest   = datetime.date.fromisoformat(_row["latest"])
        _cf_from_def = _cf_latest.replace(day=1)   # first of that month
        _cf_to_def   = _cf_latest                  # last available date
    else:
        _cf_from_def = _fy_s
        _cf_to_def   = min(_fy_e, _today)

    st.session_state["cf_from"] = _cf_from_def
    st.session_state["cf_to"]   = _cf_to_def

all_entities = ["All"] + list(ENTITIES.keys())
sel_entity   = st.sidebar.selectbox("Entity", all_entities)

all_banks = ["All"]
if sel_entity != "All":
    all_banks += ENTITIES.get(sel_entity, [])
else:
    for _bs in ENTITIES.values():
        all_banks += _bs
sel_bank = st.sidebar.selectbox("Bank Account", all_banks)

_avail_months = _cached_months(sel_fy)
month_options = ["All"] + _avail_months
sel_month     = st.sidebar.selectbox("Month", month_options)

st.sidebar.markdown("**Date Range** *(overrides Month when both set)*")
d_col1, d_col2 = st.sidebar.columns(2)
with d_col1:
    sel_date_from = st.date_input("From", value=None,
                                  format="DD/MM/YYYY", key="date_from")
with d_col2:
    sel_date_to = st.date_input("To", value=None,
                                format="DD/MM/YYYY", key="date_to")

use_date_range = sel_date_from is not None and sel_date_to is not None
date_from  = str(sel_date_from) if use_date_range else None
date_to    = str(sel_date_to)   if use_date_range else None
eff_month  = None if use_date_range else (sel_month if sel_month != "All" else None)
eff_entity = sel_entity if sel_entity != "All" else None
eff_bank   = sel_bank   if sel_bank   != "All" else None

st.sidebar.markdown("---")
st.sidebar.caption("Large debit threshold: ₹10,00,000")

if st.sidebar.button("🔄 Refresh Data", use_container_width=True):
    st.cache_data.clear()
    st.rerun()

# ── Sidebar quick stats ───────────────────────────────────────────────────────

# ─── Fetch shared data — FY filter intentionally NOT applied here ──────────────
# FY filter applies only to the Transactions tab (and the Month dropdown).
# Header KPIs, Summary, Cash Flow, Investments always show the full picture.
summary = _cached_summary(sel_entity, sel_month, None)

# B3 — bank-aware + date-range-aware totals for header KPIs
_sum_full = _cached_summary_full(
    sel_entity, sel_bank,
    sel_month if not use_date_range else None,
    str(sel_date_from) if use_date_range else None,
    str(sel_date_to)   if use_date_range else None,
)
inflow  = _sum_full["total_receipts"]
outflow = _sum_full["total_payouts"]
net     = inflow - outflow

# B4 — uncats respects all sidebar filters
uncats  = _cached_uncategorized(sel_entity, sel_bank, sel_month, sel_fy)

# ─── Closing balance for KPI card (single SQL query, no full row load) ─────────
if use_date_range:
    _cb_data = get_closing_balance(
        entity=eff_entity, bank=eff_bank,
        date_from=date_from, date_to=date_to,
    )
else:
    _cb_data = _cached_closing(sel_entity, sel_bank, sel_month, None)

by_bank_bal     = _cb_data["by_bank"]
by_bank_dt      = _cb_data["dates"]
closing_balance = _cb_data["total"]

if by_bank_bal:
    if sel_bank != "All":
        closing_date = by_bank_dt.get(sel_bank, "")
        _cb_sub      = f"As of {closing_date}" if closing_date else "No transactions"
    else:
        _dt_vals     = [v for v in by_bank_dt.values() if v]
        closing_date = max(_dt_vals) if _dt_vals else ""
        _cb_sub      = " | ".join(
            f"{bk}: ₹{v:,.0f}" for bk, v in sorted(by_bank_bal.items())
        )
else:
    closing_balance = 0.0
    closing_date    = ""
    _cb_sub         = "No transactions"

# B1 — entity balances for header KPI strip (always unfiltered / live)
stores_bal    = _cached_entity_balance("Stores")
ventures_bal  = _cached_entity_balance("Ventures")
inv_bals      = _cached_inv_bals()
stores_bank   = stores_bal
ventures_bank = ventures_bal
stores_mf     = inv_bals["Stores"]["mf_balance"]
stores_fd     = inv_bals["Stores"]["fd_balance"]
ventures_mf   = inv_bals["Ventures"]["mf_balance"]
ventures_fd   = inv_bals["Ventures"]["fd_balance"]

# ─── Cash at Stores + investment totals for header TOTAL CASH POSITION ────────
_cash_at_stores  = get_cash_at_stores()
_inv_hdr         = _cached_inv_header_totals()
_fd_total_hdr    = _inv_hdr["fd"]
_mf_total_hdr    = _inv_hdr["mf"]
_manual_totals   = get_manual_investment_totals()
_manual_inv_net  = sum((t["invested"] or 0) - (t["redeemed"] or 0) for t in _manual_totals)
_total_cash_pos  = (stores_bank + ventures_bank +
                    stores_mf + ventures_mf +
                    stores_fd + ventures_fd + _cash_at_stores)

# ─── Transaction count (lightweight SQL COUNT — no row loading) ───────────────
txn_count = _cached_count(sel_entity, sel_bank, sel_month, None)

# Period label pill
if use_date_range:
    period_label = f"{sel_date_from} → {sel_date_to}"
elif sel_month != "All":
    period_label = sel_month
else:
    period_label = "All periods"

# Entity label pill
entity_label = sel_entity if sel_entity != "All" else "All entities"

# ── Sidebar quick stats ────────────────────────────────────────────────────────
_s_receipts = inflow
_s_payments = outflow
_s_uncat    = sum(u.get("count", 0) for u in uncats)
st.sidebar.markdown('<div class="sidebar-section">Quick Stats</div>', unsafe_allow_html=True)
st.sidebar.markdown(f"""
<div class="sb-stats-grid">
  <div class="sb-stat-card">
    <div class="sb-stat-left">
      <span class="sb-stat-label">Transactions</span>
      <span class="sb-stat-value">{txn_count:,}</span>
    </div>
    <span class="sb-stat-icon">🔢</span>
  </div>
  <div class="sb-stat-card green">
    <div class="sb-stat-left">
      <span class="sb-stat-label green">Receipts</span>
      <span class="sb-stat-value green">{fmt_cr(_s_receipts)}</span>
    </div>
    <span class="sb-stat-icon">↑</span>
  </div>
  <div class="sb-stat-card red">
    <div class="sb-stat-left">
      <span class="sb-stat-label red">Payments</span>
      <span class="sb-stat-value red">{fmt_cr(_s_payments)}</span>
    </div>
    <span class="sb-stat-icon">↓</span>
  </div>
  <div class="sb-stat-card amber">
    <div class="sb-stat-left">
      <span class="sb-stat-label amber">Uncategorized</span>
      <span class="sb-stat-value amber">{_s_uncat:,}</span>
    </div>
    <span class="sb-stat-icon">⚠</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ─── Month-over-month comparison for KPI strip (from cached summary) ──────────
_this_wk_in = _this_wk_out = _wow_in = _wow_out = 0.0
_by_month = summary.get("by_month", [])
if len(_by_month) >= 2:
    _last_m      = _by_month[-1]
    _prev_m      = _by_month[-2]
    _this_wk_in  = float(_last_m.get("inflow",  0) or 0)
    _this_wk_out = float(_last_m.get("outflow", 0) or 0)
    _prev_in     = float(_prev_m.get("inflow",  0) or 0)
    _prev_out    = float(_prev_m.get("outflow", 0) or 0)
    _wow_in  = ((_this_wk_in  - _prev_in)  / _prev_in  * 100) if _prev_in  else 0.0
    _wow_out = ((_this_wk_out - _prev_out) / _prev_out * 100) if _prev_out else 0.0
elif len(_by_month) == 1:
    _this_wk_in  = float(_by_month[0].get("inflow",  0) or 0)
    _this_wk_out = float(_by_month[0].get("outflow", 0) or 0)

# ─── Header band — dark navy KPI strip ────────────────────────────────────────
net_class    = "pos" if net >= 0 else "neg"
_bank_label  = sel_bank if sel_bank != "All" else "AXIS & HDFC"
_net_arrow   = "▲" if net >= 0 else "▼"
_net_sub_cls = "kpi-positive" if net >= 0 else "kpi-negative"

# WoW badge classes and labels (pre-computed to avoid nested f-string issues)
_wow_in_cls  = "up"   if _wow_in  > 0 else ("down" if _wow_in  < 0 else "flat")
_wow_in_txt  = (f"▲ {abs(_wow_in):.1f}% WoW"  if _wow_in  > 0
               else f"▼ {abs(_wow_in):.1f}% WoW"  if _wow_in  < 0 else "— flat WoW")
_wow_out_txt = (f"▲ {abs(_wow_out):.1f}% WoW" if _wow_out > 0
               else f"▼ {abs(_wow_out):.1f}% WoW" if _wow_out < 0 else "— flat WoW")
_exc_cls     = "warn" if _s_uncat > 0 else "flat"
_exc_txt     = f"⚠ {_s_uncat:,} uncat" if _s_uncat > 0 else "✓ all categorized"
st.markdown(f"""
<div class="header-band">
    <div class="header-top">
        <div class="header-logo-area">
            <div class="header-logo-box">CK</div>
            <div>
                <div class="header-title">BankFlow Dashboard</div>
                <div class="header-subtitle">
                    {sel_entity} &nbsp;·&nbsp; {_bank_label} &nbsp;·&nbsp; {sel_fy_label} &nbsp;·&nbsp; {period_label}
                </div>
            </div>
        </div>
        <div class="header-pills">
            <span class="pill pill-live">● Live</span>
            <span class="pill pill-txn">⊞ {txn_count:,} Txns</span>
            <span class="pill pill-period">📅 {period_label}</span>
        </div>
    </div>
    <div class="kpi-strip-divider"></div>
    <div class="kpi-strip">
        <div class="kpi-strip-hero">
            <div class="kpi-strip-label">TOTAL CASH POSITION</div>
            <div style="font-size:36px;font-weight:800;color:#FFFFFF;letter-spacing:-1.2px;line-height:1;margin-bottom:3px;">{fmt_cr(_total_cash_pos)}</div>
            <div class="kpi-strip-sub kpi-neutral">Bank {fmt_cr(closing_balance)} · Cash {fmt_cr(_cash_at_stores)} · FD {fmt_cr(stores_fd+ventures_fd)} · MF {fmt_cr(stores_mf+ventures_mf)}</div>
        </div>
        <div class="kpi-strip-item">
            <div class="kpi-strip-label">STORES BANK</div>
            <div class="kpi-strip-value">{fmt_cr(stores_bank)}</div>
            <div class="kpi-strip-sub kpi-neutral">Bank closing · All accounts</div>
        </div>
        <div class="kpi-strip-item">
            <div class="kpi-strip-label">STORES MF</div>
            <div class="kpi-strip-value" style="color:#A78BFA;">{fmt_cr(stores_mf)}</div>
            <div class="kpi-strip-sub kpi-neutral">Mutual funds · Stores</div>
        </div>
        <div class="kpi-strip-item">
            <div class="kpi-strip-label">STORES FD</div>
            <div class="kpi-strip-value" style="color:#60A5FA;">{fmt_cr(stores_fd)}</div>
            <div class="kpi-strip-sub kpi-neutral">Fixed deposits · Stores</div>
        </div>
        <div class="kpi-strip-item">
            <div class="kpi-strip-label">VENTURES BANK</div>
            <div class="kpi-strip-value">{fmt_cr(ventures_bank)}</div>
            <div class="kpi-strip-sub kpi-neutral">Bank closing · All accounts</div>
        </div>
        <div class="kpi-strip-item">
            <div class="kpi-strip-label">VENTURES MF</div>
            <div class="kpi-strip-value" style="color:#A78BFA;">{fmt_cr(ventures_mf)}</div>
            <div class="kpi-strip-sub kpi-neutral">Mutual funds · Ventures</div>
        </div>
        <div class="kpi-strip-item">
            <div class="kpi-strip-label">VENTURES FD</div>
            <div class="kpi-strip-value" style="color:#60A5FA;">{fmt_cr(ventures_fd)}</div>
            <div class="kpi-strip-sub kpi-neutral">Fixed deposits · Ventures</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── Navigation ──────────────────────────────────────────────────────────────
if "active_tab" not in st.session_state:
    st.session_state["active_tab"] = "Summary"
if st.session_state.get("switch_tab"):
    st.session_state["active_tab"] = st.session_state.pop("switch_tab")

# Inject CSS to style nav buttons as tabs.
# Relies on .nav-active-indicator injected below the active button —
# that class is the reliable hook; no `key` attribute needed.
st.markdown("""
<style>
/* ── Nav bar wrapper — detected by active indicator presence ── */
div[data-testid="stHorizontalBlock"]:has(.nav-active-indicator) {
    background: #FFFFFF;
    border-bottom: 1px solid #EDF0F5;
    box-shadow: 0 1px 4px rgba(10,22,40,0.04);
    padding: 0;
    gap: 0 !important;
    margin-bottom: 16px !important;
    position: sticky;
    top: 0;
    z-index: 100;
    align-items: stretch !important;
}

/* ── All nav buttons — uniform base style ── */
div[data-testid="stHorizontalBlock"]:has(.nav-active-indicator) .stButton > button {
    background: transparent !important;
    border: none !important;
    border-radius: 0 !important;
    color: #6B7A90 !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    padding: 12px 8px 10px !important;
    width: 100% !important;
    box-shadow: none !important;
    letter-spacing: 0.01em !important;
    transition: color 0.15s ease, background 0.15s ease !important;
}
div[data-testid="stHorizontalBlock"]:has(.nav-active-indicator) .stButton > button:hover {
    background: #F4F6FA !important;
    color: #1B2B4B !important;
}

/* ── Active indicator line ── */
.nav-active-indicator {
    height: 3px;
    background: #1B2B4B;
    border-radius: 2px 2px 0 0;
    margin: -4px 8px 0;
}

/* ── Active button label ── */
div[data-testid="stHorizontalBlock"]:has(.nav-active-indicator) .stVerticalBlock:has(.nav-active-indicator) .stButton > button {
    color: #1B2B4B !important;
    font-weight: 700 !important;
}
</style>
""", unsafe_allow_html=True)

nav_options = ["Summary", "Cash Flow", "Transactions", "Uncategorised", "Exception Report", "Investments", "Upload"]
nav_keys    = ["Summary", "Cash Flow", "Transactions", "Uncategorised", "Exception Report", "Investments", "Upload"]
nav_cols = st.columns(len(nav_options))
for i, (label, key) in enumerate(zip(nav_options, nav_keys)):
    with nav_cols[i]:
        is_active = st.session_state["active_tab"] == key
        if st.button(label, key=f"nav_{key}", use_container_width=True):
            st.session_state["switch_tab"] = key
            st.rerun()
        if is_active:
            st.markdown('<div class="nav-active-indicator"></div>', unsafe_allow_html=True)

selected_tab = st.session_state["active_tab"]

# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
if selected_tab == "Summary":
    st.markdown(LIVE_BANNER, unsafe_allow_html=True)
    # ── Account Balance Summary ───────────────────────────────────────────────
    _acct_balances = get_account_balances(financial_year=sel_fy if sel_fy != "All" else None)

    def _render_entity_subtotal(entity_name, entity_bank_total, rows, _inv_bals):
        _pill_cls = "abs-stores" if entity_name == "Stores" else "abs-ventures"
        for acc in rows:
            rcols = st.columns([2, 1.5, 1.5, 1, 2])
            _vals = [
                f'<span class="abs-pill {_pill_cls}">{entity_name}</span>',
                acc["bank"],
                acc["purpose"],
                acc["bank_id"],
                f'<div style="text-align:right;font-variant-numeric:tabular-nums;">₹{acc["balance"]:,.0f}</div>',
            ]
            for _c, _v in zip(rcols, _vals):
                with _c:
                    st.markdown(_v, unsafe_allow_html=True)

        _eb  = _inv_bals.get(entity_name, {})
        _mf  = _eb.get("mf_balance", 0)
        _fd  = _eb.get("fd_balance", 0)

        if _mf:
            mrcols = st.columns([2, 1.5, 1.5, 1, 2])
            for _c, _v in zip(mrcols, [
                f'<span class="abs-pill" style="background:#EDE9FE;color:#5B21B6;">MF</span> {entity_name}',
                "Mutual Fund", "Investment", "—",
                f'<div style="text-align:right;color:#5B21B6;font-weight:700;">₹{_mf:,.0f}</div>',
            ]):
                with _c:
                    st.markdown(_v, unsafe_allow_html=True)

        if _fd:
            fdcols = st.columns([2, 1.5, 1.5, 1, 2])
            for _c, _v in zip(fdcols, [
                f'<span class="abs-pill" style="background:#DBEAFE;color:#1E40AF;">FD</span> {entity_name}',
                "Fixed Deposit", "Investment", "—",
                f'<div style="text-align:right;color:#1E40AF;font-weight:700;">₹{_fd:,.0f}</div>',
            ]):
                with _c:
                    st.markdown(_v, unsafe_allow_html=True)

        entity_total = entity_bank_total + _mf + _fd
        st.markdown(
            f'<div style="background:#F0F4FF;padding:6px 12px;border-radius:6px;'
            f'font-weight:700;color:#1B2B4B;display:flex;justify-content:space-between;'
            f'margin-bottom:8px;">'
            f'<span>{entity_name} Subtotal (Bank + MF + FD)</span>'
            f'<span>₹{entity_total:,.0f}</span></div>',
            unsafe_allow_html=True)
        return entity_total

    st.markdown('<div class="section-header" style="margin-bottom:10px;">Account Balance Summary</div>',
                unsafe_allow_html=True)

    _abs_css = """
<style>
.abs-pill{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;letter-spacing:.04em;}
.abs-stores{background:#DBEAFE;color:#1E40AF;}
.abs-ventures{background:#FCE7F3;color:#9D174D;}
</style>"""
    st.markdown(_abs_css, unsafe_allow_html=True)

    _hcols = st.columns([2, 1.5, 1.5, 1, 2])
    for _c, _hdr in zip(_hcols, ["Entity", "Bank / Type", "Purpose", "Account", "Balance"]):
        with _c:
            st.markdown(
                f'<div style="font-size:11px;font-weight:700;color:#1B2B4B;'
                f'letter-spacing:.06em;text-transform:uppercase;padding:4px 0;'
                f'border-bottom:2px solid #1B2B4B;">{_hdr}</div>',
                unsafe_allow_html=True)

    _entity_totals = {}
    for _entity_name in ["Stores", "Ventures"]:
        _entity_rows      = [r for r in _acct_balances if r["entity"] == _entity_name]
        _entity_bank_total = sum(r["balance"] for r in _entity_rows)
        _entity_totals[_entity_name] = _render_entity_subtotal(
            _entity_name, _entity_bank_total, _entity_rows, inv_bals
        )

    _grand_total = sum(_entity_totals.values()) + _cash_at_stores
    st.markdown(
        f'<div style="background:#1B2B4B;color:#fff;padding:8px 12px;'
        f'border-radius:6px;font-weight:800;font-size:14px;'
        f'display:flex;justify-content:space-between;">'
        f'<span>TOTAL CASH POSITION</span>'
        f'<span>₹{_grand_total:,.0f}</span></div>',
        unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts row ────────────────────────────────────────────────────────────
    _ch_left, _ch_right = st.columns([3, 2])

    with _ch_left:
        # Weekly treasury-style movement chart: grouped bars + net line overlay
        st.markdown('<div class="card"><div class="card-title">Monthly Cash Movement</div>',
                    unsafe_allow_html=True)
        _by_month_chart = summary.get("by_month", [])
        if _by_month_chart:
            _df_w = pd.DataFrame(_by_month_chart).tail(12).copy()
            _df_w["net"] = _df_w["inflow"] - _df_w["outflow"]
            _fig_w = go.Figure()
            _fig_w.add_trace(go.Bar(
                name="Receipts", x=_df_w["month"], y=_df_w["inflow"],
                marker=dict(color="#22C55E", opacity=0.72, line=dict(width=0)),
                hovertemplate="<b>%{x}</b><br>Receipts: ₹%{y:,.0f}<extra></extra>"
            ))
            _fig_w.add_trace(go.Bar(
                name="Payouts", x=_df_w["month"], y=_df_w["outflow"],
                marker=dict(color="#F87171", opacity=0.72, line=dict(width=0)),
                hovertemplate="<b>%{x}</b><br>Payouts: ₹%{y:,.0f}<extra></extra>"
            ))
            _fig_w.add_trace(go.Scatter(
                name="Net", x=_df_w["month"], y=_df_w["net"],
                mode="lines+markers",
                line=dict(color="#1B2B4B", width=2.5),
                marker=dict(size=5, color="#1B2B4B", symbol="circle"),
                hovertemplate="<b>%{x}</b><br>Net: ₹%{y:,.0f}<extra></extra>"
            ))
            _fig_w.update_layout(
                barmode="group", bargap=0.22, bargroupgap=0.06,
                plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                font=dict(family="Inter", color="#4A5568", size=11),
                legend=dict(
                    orientation="h", yanchor="bottom", y=1.02,
                    xanchor="right", x=1, font=dict(size=10),
                    bgcolor="rgba(0,0,0,0)"
                ),
                xaxis=dict(
                    showgrid=False, tickfont=dict(size=10),
                    tickangle=-30, linecolor="#E2E8F0", showline=True
                ),
                yaxis=dict(
                    showgrid=True, gridcolor="#F5F5F5", gridwidth=1,
                    tickfont=dict(size=10), tickformat=",.0f",
                    zeroline=True, zerolinecolor="#E2E8F0", zerolinewidth=1.5
                ),
                height=300,
                margin=dict(l=8, r=8, t=30, b=10)
            )
            st.plotly_chart(_fig_w, use_container_width=True)
        else:
            st.info("No data yet. Drop bank statements to get started.")
        st.markdown('</div>', unsafe_allow_html=True)

    with _ch_right:
        # Expenditure donut — category breakdown
        st.markdown('<div class="card"><div class="card-title">Expenditure by Category</div>',
                    unsafe_allow_html=True)
        _by_cat = summary.get("by_category", [])
        if _by_cat:
            _df_c = pd.DataFrame(_by_cat)
            _df_c = _df_c[_df_c["total"] > 0].copy()
            _df_c_top = _df_c.head(8).copy()
            if len(_df_c) > 8:
                _others = _df_c.iloc[8:]["total"].sum()
                _df_c_top = pd.concat([
                    _df_c_top,
                    pd.DataFrame([{"final_group": "Others", "total": _others}])
                ], ignore_index=True)
            _DONUT_COLORS = [
                "#1B3A6B", "#3B82F6", "#60A5FA", "#93C5FD",
                "#BFDBFE", "#1E40AF", "#2563EB", "#DBEAFE", "#E2E8F0"
            ]
            _tot_spend = _df_c["total"].sum()
            _ctr = f"₹{_tot_spend/1e7:.1f}Cr" if _tot_spend >= 1e7 else f"₹{_tot_spend/1e5:.1f}L"
            _fig_c = go.Figure(go.Pie(
                labels=_df_c_top["final_group"],
                values=_df_c_top["total"],
                hole=0.62,
                marker=dict(
                    colors=_DONUT_COLORS[:len(_df_c_top)],
                    line=dict(color="#FFFFFF", width=2)
                ),
                textinfo="none",
                hovertemplate="<b>%{label}</b><br>₹%{value:,.0f}<br>%{percent}<extra></extra>"
            ))
            _fig_c.add_annotation(
                text=f"{_ctr}<br><span style='font-size:9px'>Total Spend</span>",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=14, color="#1A202C", family="Inter"),
                align="center"
            )
            _fig_c.update_layout(
                plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                font=dict(family="Inter", color="#4A5568", size=11),
                legend=dict(
                    orientation="v", yanchor="middle", y=0.5,
                    xanchor="left", x=1.02, font=dict(size=10),
                    itemsizing="constant"
                ),
                height=300,
                margin=dict(l=8, r=110, t=30, b=10),
                showlegend=True
            )
            st.plotly_chart(_fig_c, use_container_width=True)
        else:
            st.info("No categorized data yet.")
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Analysis charts (15-month trend, YoY, top expenses, weekly) ─────────────
    import datetime as _dt
    _cur_month = _latest_date[:7] if _latest_date else _dt.date.today().strftime("%Y-%m")

    # Chart 1 — 15-month receipts vs payouts trend (full width)
    _trend = get_monthly_trend(15)
    if _trend:
        _df_t = pd.DataFrame(_trend)
        _fig_t = go.Figure()
        _fig_t.add_trace(go.Bar(
            name="Receipts", x=_df_t["month"],
            y=_df_t["receipts"] / 1e7,
            marker_color="#22C55E", opacity=0.8,
            hovertemplate="%{x}<br>₹%{y:.2f} Cr<extra>Receipts</extra>"
        ))
        _fig_t.add_trace(go.Bar(
            name="Payouts", x=_df_t["month"],
            y=_df_t["payouts"] / 1e7,
            marker_color="#F87171", opacity=0.8,
            hovertemplate="%{x}<br>₹%{y:.2f} Cr<extra>Payouts</extra>"
        ))
        _fig_t.add_trace(go.Scatter(
            name="Net", x=_df_t["month"],
            y=(_df_t["receipts"] - _df_t["payouts"]) / 1e7,
            mode="lines+markers",
            line=dict(color="#1B2B4B", width=2),
            hovertemplate="%{x}<br>Net ₹%{y:.2f} Cr<extra></extra>"
        ))
        _fig_t.update_layout(
            title="15-Month Receipts vs Payouts (₹ Cr)",
            barmode="group", height=300,
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Inter", size=11),
            legend=dict(orientation="h", y=1.1),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#F0F0F0", ticksuffix=" Cr"),
            margin=dict(l=0, r=0, t=40, b=0)
        )
        st.plotly_chart(_fig_t, use_container_width=True)

    # Charts 2 + 3 — YoY comparison and Top expenses (side by side)
    _c1, _c2 = st.columns(2)

    with _c1:
        _yoy = get_yoy_comparison(_cur_month)
        _cy  = _yoy["current"]["month"][:4]
        _ly  = _yoy["last_year"]["month"][:4]
        _mo_name = _dt.date(int(_cur_month[:4]), int(_cur_month[5:]), 1).strftime("%B")
        _fig_y = go.Figure()
        for _lbl, _cv, _lv, _col in [
            ("Receipts", _yoy["current"]["receipts"] / 1e7,
             _yoy["last_year"]["receipts"] / 1e7, "#22C55E"),
            ("Payouts",  _yoy["current"]["payouts"]  / 1e7,
             _yoy["last_year"]["payouts"]  / 1e7, "#F87171"),
        ]:
            _fig_y.add_trace(go.Bar(
                name=f"{_lbl} {_cy}", x=[_lbl], y=[_cv],
                marker_color=_col, opacity=0.9))
            _fig_y.add_trace(go.Bar(
                name=f"{_lbl} {_ly}", x=[_lbl], y=[_lv],
                marker_color=_col, opacity=0.4))
        _fig_y.update_layout(
            title=f"YoY — {_mo_name} {_ly} vs {_cy} (₹ Cr)",
            barmode="group", height=300,
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Inter", size=11),
            legend=dict(orientation="h", y=1.1),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#F0F0F0", ticksuffix=" Cr"),
            margin=dict(l=0, r=0, t=40, b=0)
        )
        st.plotly_chart(_fig_y, use_container_width=True)
        for _lbl, _cv, _lv in [
            ("Receipts", _yoy["current"]["receipts"], _yoy["last_year"]["receipts"]),
            ("Payouts",  _yoy["current"]["payouts"],  _yoy["last_year"]["payouts"]),
        ]:
            _chg = ((_cv - _lv) / _lv * 100) if _lv else 0
            _pcol = "#16A34A" if _chg >= 0 else "#DC2626"
            _arr  = "▲" if _chg >= 0 else "▼"
            st.markdown(
                f'<span style="background:#F5F5F5; padding:3px 10px; '
                f'border-radius:10px; font-size:12px; margin-right:8px;">'
                f'{_lbl}: <b style="color:{_pcol};">{_arr} {abs(_chg):.1f}%</b>'
                f'</span>', unsafe_allow_html=True)

    with _c2:
        _exp = get_top_expenses_comparison(_cur_month)
        if _exp["categories"]:
            _pm = _dt.date(int(_exp["prev_month"][:4]),
                           int(_exp["prev_month"][5:]), 1).strftime("%b %Y")
            _cm = _dt.date(int(_exp["current_month"][:4]),
                           int(_exp["current_month"][5:]), 1).strftime("%b %Y")
            _fig_e = go.Figure()
            _fig_e.add_trace(go.Bar(
                name=_cm, y=_exp["categories"],
                x=[v / 1e5 for v in _exp["current"]],
                orientation="h", marker_color="#1B2B4B", opacity=0.85,
                hovertemplate="%{y}<br>₹%{x:.1f} L<extra>This month</extra>"
            ))
            _fig_e.add_trace(go.Bar(
                name=_pm, y=_exp["categories"],
                x=[v / 1e5 for v in _exp["previous"]],
                orientation="h", marker_color="#93C5FD", opacity=0.7,
                hovertemplate="%{y}<br>₹%{x:.1f} L<extra>Last month</extra>"
            ))
            _fig_e.update_layout(
                title=f"Top Expenses: {_pm} vs {_cm} (₹ L)",
                barmode="group", height=300,
                plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
                font=dict(family="Inter", size=11),
                legend=dict(orientation="h", y=1.1),
                xaxis=dict(showgrid=True, gridcolor="#F0F0F0", ticksuffix=" L"),
                yaxis=dict(showgrid=False, categoryorder="total ascending"),
                margin=dict(l=0, r=80, t=40, b=0)
            )
            st.plotly_chart(_fig_e, use_container_width=True)

    # Chart 4 — 12-week net cash flow area chart (full width)
    _wkp = get_weekly_cash_position(12)
    if _wkp:
        _df_wk = pd.DataFrame(_wkp)
        _df_wk["label"] = _df_wk["week_start"].apply(
            lambda d: _dt.date.fromisoformat(d).strftime("%d %b") if d else "")
        _fig_wk = go.Figure()
        _fig_wk.add_trace(go.Scatter(
            x=_df_wk["label"],
            y=_df_wk["net_flow"] / 1e7,
            mode="lines+markers",
            fill="tozeroy",
            line=dict(color="#1B2B4B", width=2),
            marker=dict(size=5),
            fillcolor="rgba(27,43,75,0.08)",
            hovertemplate="Week of %{x}<br>Net ₹%{y:.2f} Cr<extra></extra>"
        ))
        _fig_wk.add_hline(y=0, line_dash="dash",
                          line_color="#E53E3E", line_width=1)
        _fig_wk.update_layout(
            title="12-Week Net Cash Flow (₹ Cr)",
            height=280,
            plot_bgcolor="#FFFFFF", paper_bgcolor="#FFFFFF",
            font=dict(family="Inter", size=11),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#F0F0F0", ticksuffix=" Cr"),
            margin=dict(l=0, r=0, t=40, b=0)
        )
        st.plotly_chart(_fig_wk, use_container_width=True)

    # ── Monthly trend + category detail (full width) ──────────────────────────
    render_overview(summary)

# ══════════════════════════════════════════════════════════════════════════════
# TRANSACTIONS
# ══════════════════════════════════════════════════════════════════════════════
elif selected_tab == "Transactions":
    # ── Category/Group options from DB — never loads all rows ─────────────────
    all_cats_for_txn = get_all_categories()
    fg_options = (
        ["All", "🔒 Manually Categorized", "Uncategorized"]
        + sorted([r["final_group"] for r in all_cats_for_txn])
    )
    mg_options = ["All"] + sorted({
        r["main_group"] for r in all_cats_for_txn if r.get("main_group")
    })

    # ── Clear filter flag — checked before widgets ─────────────────────────────
    if st.session_state.get("clear_txn_filters"):
        st.session_state["clear_txn_filters"] = False
        st.session_state["txn_fg"]     = "All"
        st.session_state["txn_mg"]     = "All"
        st.session_state["txn_search"] = ""

    # ── Read filter values from session state (widgets rendered below OB table) ─
    sel_fg = st.session_state.get("txn_fg", "All")
    sel_mg = st.session_state.get("txn_mg", "All")
    search = st.session_state.get("txn_search", "")

    # ── Map UI selections to SQL-level params ─────────────────────────────────
    _pag_cat     = None
    _pag_manonly = False
    if sel_fg == "🔒 Manually Categorized":
        _pag_manonly = True
    elif sel_fg == "Uncategorized":
        _pag_cat = "Uncategorized"
    elif sel_fg != "All":
        _pag_cat = sel_fg
    _pag_mg = sel_mg if sel_mg != "All" else None

    # ── Pagination state — reset when any filter changes ──────────────────────
    if "txn_page" not in st.session_state:
        st.session_state["txn_page"] = 1
    _fk = f"{sel_entity}_{sel_bank}_{sel_month}_{sel_fy}_{sel_fg}_{_pag_mg}_{search}"
    if st.session_state.get("_last_filter_key") != _fk:
        st.session_state["txn_page"] = 1
        st.session_state["_last_filter_key"] = _fk

    PAGE_SIZE = 500

    # ── Opening Balance rows (≤1 per bank — tiny separate query) ──────────────
    _ob_raw = get_all_transactions(
        entity=eff_entity, bank=eff_bank,
        month=eff_month, date_from=date_from, date_to=date_to,
        financial_year=sel_fy, category="OPENING BALANCE",
    )
    ob_rows = pd.DataFrame(_ob_raw) if _ob_raw else pd.DataFrame()

    # ── Paginated main rows ────────────────────────────────────────────────────
    pag = get_paginated_transactions(
        entity=eff_entity, bank=eff_bank,
        month=eff_month, date_from=date_from, date_to=date_to,
        category=_pag_cat,
        main_group=_pag_mg,
        manually_overridden_only=_pag_manonly,
        financial_year=sel_fy,
        search=search if search else None,
        page=st.session_state["txn_page"],
        page_size=PAGE_SIZE,
    )
    txn_rows     = pd.DataFrame(pag["rows"]) if pag["rows"] else pd.DataFrame()
    _total_count = pag["total_count"]
    _total_pages = pag["total_pages"]
    _cur_page    = pag["page"]

    if _total_count > 0 or not ob_rows.empty:
        _ob_suffix = f" + {len(ob_rows)} OB row(s)" if not ob_rows.empty else ""
        st.markdown(f"""
<span class="txn-count-badge">
  <span class="txn-dot"></span>
  {_total_count:,} transactions{_ob_suffix}
</span>
""", unsafe_allow_html=True)

        display_cols = ["date", "entity", "bank", "narration", "debit", "credit",
                        "balance", "final_group", "group_name", "main_group"]

        if not ob_rows.empty:
            st.markdown(
                '<div class="ob-banner">'
                '<span class="ob-banner-label">Opening Balance</span>'
                '</div>',
                unsafe_allow_html=True
            )
            ob_show = ob_rows[[c for c in display_cols if c in ob_rows.columns]].copy()
            for _c in ["debit", "credit", "balance"]:
                if _c in ob_show.columns:
                    ob_show[_c] = ob_show[_c].apply(fmt_inr)
            ob_show.columns = [c.replace("_", " ").title() for c in ob_show.columns]
            st.dataframe(ob_show, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("""
<div class="txn-filter-bar">
  <div class="txn-filter-title">Transaction Filters</div>
""", unsafe_allow_html=True)
        _fc1, _fc2, _fc3, _fc4 = st.columns([2, 2, 2, 1])
        with _fc1:
            sel_fg = st.selectbox(
                "Category", fg_options,
                index=fg_options.index(st.session_state.get("txn_fg", "All"))
                      if st.session_state.get("txn_fg", "All") in fg_options else 0,
                key="txn_fg")
        with _fc2:
            sel_mg = st.selectbox(
                "Main Group", mg_options,
                index=mg_options.index(st.session_state.get("txn_mg", "All"))
                      if st.session_state.get("txn_mg", "All") in mg_options else 0,
                key="txn_mg")
        with _fc3:
            search = st.text_input("Search narration", key="txn_search",
                                   placeholder="Type to search…")
        with _fc4:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✕ Clear", use_container_width=True):
                st.session_state["clear_txn_filters"] = True
                try:
                    st.rerun()
                except AttributeError:
                    st.experimental_rerun()
        st.markdown('</div>', unsafe_allow_html=True)

        if not txn_rows.empty:
            # ── Category lookup ───────────────────────────────────────────────
            all_cats   = all_cats_for_txn
            cat_names  = [r["final_group"] for r in all_cats]
            cat_lookup = {r["final_group"]: r for r in all_cats}
            cat_options = ["-- Select Category --"] + cat_names

            # ── Categorization panel ──────────────────────────────────────────
            st.markdown("""
<div class="txn-cat-panel">
  <div class="txn-cat-title">Bulk Categorization</div>
""", unsafe_allow_html=True)
            p1, p2, p3, p4 = st.columns([3, 2, 2, 2])
            with p1:
                sel_cat = st.selectbox(
                    "Final Group", cat_options, key="bulk_cat_select", index=0)

            # Auto-fill group name — only reset when category selection changes
            if st.session_state.get("_prev_cat") != sel_cat:
                st.session_state["_auto_group"] = (
                    cat_lookup.get(sel_cat, {}).get("group_name", "")
                    if sel_cat != "-- Select Category --" else ""
                )
                st.session_state["_prev_cat"] = sel_cat

            with p2:
                group_name_val = st.text_input(
                    "Group Name",
                    value=st.session_state.get("_auto_group", ""),
                    key="manual_group_input")
                st.session_state["_auto_group"] = group_name_val

            with p3:
                st.markdown("<br>", unsafe_allow_html=True)
                apply_btn = st.button(
                    "✅ Apply to Selected",
                    key="apply_bulk_cat",
                    use_container_width=True)

            with p4:
                st.markdown("<br>", unsafe_allow_html=True)
                revert_btn = st.button(
                    "↩️ Revert Selected",
                    key="revert_bulk_btn",
                    use_container_width=True)

            # ── Select All / Deselect All ─────────────────────────────────────
            if "select_all_rows" not in st.session_state:
                st.session_state["select_all_rows"] = None

            _sa1, _sa2, _ = st.columns([1, 1, 5])
            with _sa1:
                if st.button("☑ Select All", key="select_all_btn",
                             use_container_width=True):
                    st.session_state["select_all_rows"] = True
                    st.rerun()
            with _sa2:
                if st.button("☐ Deselect All", key="deselect_all_btn",
                             use_container_width=True):
                    st.session_state["select_all_rows"] = False
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)

            # ── Data editor ───────────────────────────────────────────────────
            df_editor = txn_rows[["id", "date", "entity", "bank", "narration",
                                   "debit", "credit", "balance",
                                   "final_group", "group_name", "main_group",
                                   "manually_overridden"]].copy()
            df_editor.insert(0, "Select", False)

            if st.session_state["select_all_rows"] is True:
                df_editor["Select"] = True
            elif st.session_state["select_all_rows"] is False:
                df_editor["Select"] = False

            df_editor["debit"]   = df_editor["debit"].apply(
                lambda x: fmt_inr(float(x)) if float(x or 0) > 0 else "—")
            df_editor["credit"]  = df_editor["credit"].apply(
                lambda x: fmt_inr(float(x)) if float(x or 0) > 0 else "—")
            df_editor["balance"] = df_editor["balance"].apply(
                lambda x: fmt_inr(float(x)) if x and str(x) not in ("None", "nan") else "—")
            df_editor["manually_overridden"] = df_editor["manually_overridden"].apply(
                lambda x: "🔒" if int(x or 0) == 1 else "")

            edited = st.data_editor(
                df_editor,
                column_config={
                    "Select":              st.column_config.CheckboxColumn("✓", width="small"),
                    "id":                  st.column_config.NumberColumn("ID", width="small"),
                    "date":                st.column_config.Column("Date", width="small"),
                    "entity":              st.column_config.Column("Entity", width="small"),
                    "bank":                st.column_config.Column("Bank", width="small"),
                    "narration":           st.column_config.Column("Narration", width="large"),
                    "debit":               st.column_config.Column("Debit ↓", width="medium"),
                    "credit":              st.column_config.Column("Credit ↑", width="medium"),
                    "balance":             st.column_config.Column("Balance", width="medium"),
                    "final_group":         st.column_config.Column("Category", width="medium"),
                    "group_name":          st.column_config.Column("Group", width="medium"),
                    "main_group":          st.column_config.Column("Main Group", width="small"),
                    "manually_overridden": st.column_config.Column("🔒", width="small"),
                },
                disabled=["id", "date", "entity", "bank", "narration",
                          "debit", "credit", "balance",
                          "final_group", "group_name", "main_group",
                          "manually_overridden"],
                hide_index=True,
                use_container_width=True,
                height=560,
                key="txn_data_editor"
            )

            # ── Apply / Revert handlers ───────────────────────────────────────
            def _get_selected_ids(edited_df):
                sel = edited_df[edited_df["Select"] == True]
                if sel.empty:
                    return []
                return [int(i) for i in sel["id"].tolist() if i]

            if apply_btn:
                if sel_cat == "-- Select Category --":
                    st.warning("⚠️ Select a Final Group first.")
                else:
                    selected_ids = _get_selected_ids(edited)
                    if not selected_ids:
                        st.warning("⚠️ No rows selected.")
                    else:
                        receipt_ids, payment_ids = [], []
                        for txn_id in selected_ids:
                            orig = txn_rows[txn_rows["id"] == txn_id]
                            if not orig.empty:
                                cr = float(orig.iloc[0].get("credit", 0) or 0)
                                receipt_ids.append(txn_id) if cr > 0 \
                                    else payment_ids.append(txn_id)
                        total = 0
                        if receipt_ids:
                            total += manual_categorize(receipt_ids, sel_cat,
                                                       group_name_val, "Receipt")
                        if payment_ids:
                            total += manual_categorize(payment_ids, sel_cat,
                                                       group_name_val, "Payment")
                        st.cache_data.clear()
                        st.success(f"✅ {total} transaction(s) categorized as '{sel_cat}'")
                        st.session_state["select_all_rows"] = None
                        st.rerun()

            if revert_btn:
                selected_ids = _get_selected_ids(edited)
                if not selected_ids:
                    st.warning("⚠️ No rows selected.")
                else:
                    reverted   = sum(1 for i in selected_ids if revert_manual_category(i))
                    no_history = len(selected_ids) - reverted
                    if reverted:
                        st.success(f"↩️ Reverted {reverted} transaction(s) to previous category.")
                    if no_history:
                        st.warning(f"{no_history} row(s) had no previous category to revert to.")
                    st.session_state["select_all_rows"] = None
                    st.rerun()

        # ── Legend ────────────────────────────────────────────────────────────
        st.markdown("""
<div class="legend-strip">
  <span class="legend-badge green">🔒 Manually set — protected from Reload</span>
  <span class="legend-badge amber">Uncategorized rows need attention</span>
</div>
""", unsafe_allow_html=True)

        # ── Audit History ─────────────────────────────────────────────────────
        st.markdown("---")
        with st.expander("🕐 Category Change History", expanded=False):
            audit_log = get_category_audit(limit=50)
            if audit_log:
                df_audit = pd.DataFrame(audit_log)[[
                    "changed_at", "transaction_id", "bank", "entity",
                    "narration", "old_category", "new_category", "change_type"
                ]]
                df_audit.columns = [
                    "Changed At", "Txn ID", "Bank", "Entity",
                    "Narration", "From", "To", "Type"
                ]
                def _color_type(val):
                    return "color:#D97706" if val == "revert" else "color:#059669"
                st.dataframe(
                    df_audit.style.map(_color_type, subset=["Type"]),
                    use_container_width=True, hide_index=True, height=300)
                st.download_button(
                    "⬇️ Export Audit Log",
                    df_audit.to_csv(index=False).encode("utf-8"),
                    "category_audit.csv", "text/csv")
            else:
                st.info("No category changes recorded yet.")

        # ── Pagination controls ───────────────────────────────────────────────
        if _total_pages > 1:
            pg1, pg2, pg3, pg4, pg5 = st.columns([1, 1, 2, 1, 1])
            with pg1:
                if st.button("⟨⟨ First", key="txn_first_pg",
                             disabled=_cur_page == 1):
                    st.session_state["txn_page"] = 1; st.rerun()
            with pg2:
                if st.button("⟨ Prev", key="txn_prev_pg",
                             disabled=_cur_page == 1):
                    st.session_state["txn_page"] -= 1; st.rerun()
            with pg3:
                st.markdown(
                    f'<div style="text-align:center;padding:6px;font-size:13px;">'
                    f'Page {_cur_page} of {_total_pages} '
                    f'({_total_count:,} transactions)</div>',
                    unsafe_allow_html=True)
            with pg4:
                if st.button("Next ⟩", key="txn_next_pg",
                             disabled=_cur_page == _total_pages):
                    st.session_state["txn_page"] += 1; st.rerun()
            with pg5:
                if st.button("Last ⟩⟩", key="txn_last_pg",
                             disabled=_cur_page == _total_pages):
                    st.session_state["txn_page"] = _total_pages; st.rerun()

        # ── Reconciliation summary strip (uses cached summary totals) ─────────
        _ob_bal   = float(ob_rows["balance"].iloc[0]) if not ob_rows.empty else None
        _total_cr = inflow
        _total_dr = outflow
        _movement = _total_cr - _total_dr
        _mv_cls   = "pos" if _movement >= 0 else "neg"
        st.markdown(f"""
<div class="cf-stmt-band" style="margin-top:14px;">
  <div class="cf-stmt-band-kpi">
    <span class="cf-stmt-band-label">Opening Balance</span>
    <span class="cf-stmt-band-val">{fmt_inr(_ob_bal)}</span>
  </div>
  <div class="cf-stmt-band-kpi">
    <span class="cf-stmt-band-label">Total Credits</span>
    <span class="cf-stmt-band-val pos">{fmt_inr(_total_cr)}</span>
  </div>
  <div class="cf-stmt-band-kpi">
    <span class="cf-stmt-band-label">Total Debits</span>
    <span class="cf-stmt-band-val neg">{fmt_inr(_total_dr)}</span>
  </div>
  <div class="cf-stmt-band-kpi">
    <span class="cf-stmt-band-label">Net Movement</span>
    <span class="cf-stmt-band-val {_mv_cls}">{fmt_inr(_movement)}</span>
  </div>
  <div class="cf-stmt-band-kpi">
    <span class="cf-stmt-band-label">Closing Balance</span>
    <span class="cf-stmt-band-val">{fmt_inr(closing_balance)}</span>
  </div>
</div>""", unsafe_allow_html=True)

        if not txn_rows.empty:
            _csv_fname = f"transactions_{sel_fg}_{_pag_mg or 'All'}.csv".replace(" ", "_")
            st.download_button(
                label=f"⬇️ Download page ({len(txn_rows):,} rows) as CSV",
                data=txn_rows.to_csv(index=False).encode("utf-8"),
                file_name=_csv_fname,
                mime="text/csv"
            )
    else:
        st.info("No transactions found. Adjust filters or drop bank statement files.")

# ══════════════════════════════════════════════════════════════════════════════
# REVIEW QUEUE
# ══════════════════════════════════════════════════════════════════════════════
elif selected_tab == "Uncategorised":
    render_review_queue(uncats)

# ══════════════════════════════════════════════════════════════════════════════
# EXCEPTIONS
# ══════════════════════════════════════════════════════════════════════════════
elif selected_tab == "Exception Report":
    st.markdown("### Exceptions & Alerts")
    st.markdown("""
<div style="background:#EBF4FF; border-left:4px solid #3B82F6;
     border-radius:6px; padding:10px 16px; margin-bottom:16px;
     font-size:13px; color:#1E40AF;">
    📡 <b>Live view</b> — This tab always shows the full picture
    across all entities and periods, independent of sidebar filters.
</div>
""", unsafe_allow_html=True)

    # All calls in this tab are unfiltered — full DB picture regardless of sidebar
    _exc_uncats      = get_uncategorized()
    _exc_stmt_status = get_missing_statements()
    _exc_large_debits = get_large_debits(LARGE_DEBIT_THRESHOLD)

    uncat_count = sum(u["count"] for u in _exc_uncats)
    if uncat_count > 0:
        by_entity_unc = {}
        for u in _exc_uncats:
            e = u.get("entity", "Unknown")
            by_entity_unc[e] = by_entity_unc.get(e, 0) + u["count"]
        breakdown      = " | ".join([f"{e}: {c}" for e, c in by_entity_unc.items()])
        top_narrations = ", ".join([u["narration"][:40] for u in _exc_uncats[:3]])
        st.markdown(f"""
        <div class="alert-box alert-red">
            <b>🔴 UNCATEGORIZED TRANSACTIONS</b><br>
            <b>{uncat_count}</b> transactions need categorization<br>
            <small>{breakdown}</small><br>
            <small>Top unmatched: {top_narrations}</small>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="alert-box alert-green">
            <b>✅ UNCATEGORIZED</b> — All transactions are categorized.
        </div>""", unsafe_allow_html=True)

    st.markdown("")

    st.markdown("#### 🏦 Bank Statement Status")
    for s in _exc_stmt_status:
        if s["status"] == "ok":
            bg, border, icon = "#F0FFF4", "#38A169", "✅"
            msg = f"Current — last entry <b>{s['last_date']}</b>"
        elif s["status"] == "warning":
            bg, border, icon = "#FFFFF0", "#D69E2E", "⚠️"
            msg = (f"Last loaded: <b>{s['last_date']}</b> "
                   f"({s['days_ago']} days ago) — "
                   f"current month not loaded")
        else:
            bg, border, icon = "#FFF5F5", "#E53E3E", "🔴"
            msg = f"No data loaded — {s['last_date']}"

        _cb = s.get("closing_bal")
        _cb_str = (f" · Closing: ₹{fmt_inr(_cb)}" if _cb is not None else "")
        st.markdown(f"""
    <div style="background:{bg}; border-left:4px solid {border};
         border-radius:6px; padding:10px 16px; margin:5px 0;
         display:flex; justify-content:space-between;">
        <span style="font-size:13px; color:#1A202C;">
            {icon}&nbsp; <b>{s['entity']} — {s['bank']}</b>
            &nbsp;|&nbsp; {msg}
        </span>
        <span style="font-size:11px; color:#8896A5;">
            {s['total_rows']:,} rows{_cb_str}
        </span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("")

    # ── Upload Trail ──────────────────────────────────────────────────────────
    st.markdown("#### Upload Trail")
    st.caption("History of all loaded statements. Click Delete to remove a wrongly uploaded file.")
    _trail = get_upload_trail()
    if not _trail:
        st.info("No statements uploaded yet.")
    else:
        _th = st.columns([2, 1, 1, 1, 1, 1, 0.8])
        for _tc, _tl in zip(_th, ["Filename", "Entity", "Bank", "FY",
                                    "Rows", "Uploaded At", "Action"]):
            with _tc:
                st.markdown(
                    f'<div style="font-size:11px;font-weight:700;color:#8896A5;'
                    f'text-transform:uppercase;">{_tl}</div>',
                    unsafe_allow_html=True)
        st.markdown('<hr style="margin:4px 0;border:none;border-top:1px solid #E8ECF0;">',
                    unsafe_allow_html=True)
        for _t in _trail:
            _tc = st.columns([2, 1, 1, 1, 1, 1, 0.8])
            with _tc[0]:
                st.markdown(f'<div style="font-size:12px;color:#1A202C;padding:4px 0;'
                            f'word-break:break-all;">{_t["filename"]}</div>',
                            unsafe_allow_html=True)
            with _tc[1]:
                st.markdown(f'<div style="font-size:12px;color:#4A5568;padding:4px 0;">'
                            f'{_t["entity"] or "—"}</div>', unsafe_allow_html=True)
            with _tc[2]:
                st.markdown(f'<div style="font-size:12px;color:#4A5568;padding:4px 0;">'
                            f'{_t["bank"] or "—"}</div>', unsafe_allow_html=True)
            with _tc[3]:
                st.markdown(f'<div style="font-size:12px;color:#4A5568;padding:4px 0;">'
                            f'{_t["financial_year"] or "—"}</div>', unsafe_allow_html=True)
            with _tc[4]:
                st.markdown(f'<div style="font-size:12px;color:#4A5568;padding:4px 0;">'
                            f'{_t["rows_inserted"]:,}</div>', unsafe_allow_html=True)
            with _tc[5]:
                st.markdown(f'<div style="font-size:12px;color:#8896A5;padding:4px 0;">'
                            f'{_t["uploaded_at"][:16]}</div>', unsafe_allow_html=True)
            with _tc[6]:
                if st.button("Delete", key=f"del_upload_{_t['id']}"):
                    st.session_state[f"confirm_del_{_t['id']}"] = True
                    st.rerun()
            if st.session_state.get(f"confirm_del_{_t['id']}"):
                st.warning(f"Delete **{_t['filename']}** ({_t['rows_inserted']:,} rows)? "
                           f"This cannot be undone.")
                _dc1, _dc2 = st.columns([1, 4])
                with _dc1:
                    if st.button("Confirm Delete", key=f"confirm_yes_{_t['id']}"):
                        _deleted = delete_upload(_t["id"])
                        st.success(f"Deleted {_deleted:,} rows from {_t['filename']}")
                        del st.session_state[f"confirm_del_{_t['id']}"]
                        st.cache_data.clear()
                        st.rerun()
                with _dc2:
                    if st.button("Cancel", key=f"confirm_no_{_t['id']}"):
                        del st.session_state[f"confirm_del_{_t['id']}"]
                        st.rerun()
            st.markdown('<hr style="margin:2px 0;border:none;border-top:1px solid #F0F0F0;">',
                        unsafe_allow_html=True)

    st.markdown("")

    # ── Transfer Reconciliation ───────────────────────────────────────────────
    recon  = get_transfer_reconciliation()
    ib_net = recon["interbank_net"]
    ic_net = recon["intercompany_net"]

    st.markdown("#### 🔄 Transfer Reconciliation")
    for _lbl, _net, _warn_bg, _warn_txt, _hint in [
        ("Interbank",    ib_net, "#FFFFF0", "#744210",
         "Check for missing entries across bank accounts."),
        ("Intercompany", ic_net, "#FFF5F5", "#742A2A",
         "Stores vs Ventures transfers do not balance."),
    ]:
        if _net is not None:
            _direction = "Receipt side higher" if _net > 0 else "Payout side higher"
            st.markdown(f"""
        <div style="background:{_warn_bg}; border-left:4px solid #D69E2E;
             border-radius:6px; padding:10px 16px; margin:6px 0;
             font-size:13px; color:{_warn_txt};">
            ⚠️ <b>{_lbl} difference: {fmt_inr(abs(_net))}</b>
            — {_direction}. {_hint}
        </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
        <div style="background:#F0FFF4; border-left:4px solid #38A169;
             border-radius:6px; padding:10px 16px; margin:6px 0;
             font-size:13px; color:#276749;">
            ✅ <b>{_lbl} transfers fully netted</b> — no difference.
        </div>""", unsafe_allow_html=True)

    st.markdown("")

    large_debits = _exc_large_debits
    if large_debits:
        rows_html = "".join([
            f"<tr>"
            f"<td style='padding:5px 8px'>{r['date']}</td>"
            f"<td style='padding:5px 8px'>{r['entity']}</td>"
            f"<td style='padding:5px 8px'>{r['bank']}</td>"
            f"<td style='padding:5px 8px;color:var(--red-700,#B91C1C);font-weight:600'>"
            f"₹{r['debit']:,.0f}</td>"
            f"<td style='padding:5px 8px'>{r['narration'][:50]}</td>"
            f"<td style='padding:5px 8px;color:var(--tx-3,#6B7A90)'>{r['final_group'] or '—'}</td>"
            f"</tr>"
            for r in large_debits[:10]
        ])
        st.markdown(f"""
        <div class="alert-box alert-yellow">
            <b>🟡 LARGE DEBITS</b> (above ₹{LARGE_DEBIT_THRESHOLD:,.0f})<br><br>
            <table style="width:100%;color:var(--tx-2,#3D4F66);font-size:13px;
                          border-collapse:collapse;font-family:'Inter',sans-serif;">
                <tr style="color:var(--tx-4,#96A3B4);border-bottom:1px solid var(--border-sm,#E2E7EF);
                           font-size:9.5px;font-weight:700;letter-spacing:0.09em;text-transform:uppercase;">
                    <th style="padding:6px 8px;text-align:left">Date</th>
                    <th style="padding:6px 8px;text-align:left">Entity</th>
                    <th style="padding:6px 8px;text-align:left">Bank</th>
                    <th style="padding:6px 8px;text-align:left">Amount</th>
                    <th style="padding:6px 8px;text-align:left">Narration</th>
                    <th style="padding:6px 8px;text-align:left">Category</th>
                </tr>
                {rows_html}
            </table>
        </div>""", unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="alert-box alert-green">
            <b>✅ LARGE DEBITS</b> — No debits above ₹{LARGE_DEBIT_THRESHOLD:,.0f}
        </div>""", unsafe_allow_html=True)

    st.markdown("")

    import sqlite3 as _sq3_exc
    _exc_conn = _sq3_exc.connect(DATABASE_FILE)
    _exc_conn.row_factory = _sq3_exc.Row
    _inter_row = _exc_conn.execute("""
        SELECT COUNT(*) as cnt, ROUND(SUM(debit),2) as total
        FROM transactions
        WHERE main_group IN ('INTERBANK','INTERCOMPANY')
    """).fetchone()
    _exc_conn.close()
    if _inter_row and _inter_row["cnt"] > 0:
        st.markdown(f"""
        <div class="alert-box alert-green">
            <b>ℹ️ INTERBANK / INTERCOMPANY TRANSFERS</b> — Excluded from P&amp;L<br>
            {_inter_row["cnt"]:,} transfers totalling ₹{(_inter_row["total"] or 0):,.0f} are excluded
            from inflow/outflow calculations.
        </div>""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# CASH FLOW
# ══════════════════════════════════════════════════════════════════════════════
elif selected_tab == "Cash Flow":
    st.markdown(LIVE_BANNER, unsafe_allow_html=True)
    # ── Toggle session state ──────────────────────────────────────────────────
    for _k in ["cf_ob_open", "cf_rec_open", "cf_pay_open", "cf_cb_open"]:
        if _k not in st.session_state:
            st.session_state[_k] = False
    for _k in ["wcf_ob_open", "wcf_rec_open", "wcf_pay_open", "wcf_cb_open"]:
        if _k not in st.session_state:
            st.session_state[_k] = False

    # ── Styled filter bar ─────────────────────────────────────────────────────
    st.markdown("""
<div class="cf-control-bar">
  <div class="cf-bar-label">Cash Flow Statement</div>
""", unsafe_allow_html=True)

    _cc1, _cc2, _cc3, _cc4 = st.columns([2, 2, 1.5, 1.5])
    with _cc1:
        st.markdown('<p style="color:rgba(255,255,255,0.6);font-size:13px;font-weight:600;margin-bottom:2px;">ENTITY</p>', unsafe_allow_html=True)
        cf_ent = st.radio("", ["All", "Stores", "Ventures"], horizontal=True, key="cf_ent_radio", label_visibility="collapsed")
    with _cc2:
        st.markdown('<p style="color:rgba(255,255,255,0.6);font-size:13px;font-weight:600;margin-bottom:2px;">VIEW</p>', unsafe_allow_html=True)
        cf_view = st.radio("", ["Monthly", "Weekly"], horizontal=True, key="cf_view_radio", label_visibility="collapsed")
    with _cc3:
        st.markdown('<p style="color:rgba(255,255,255,0.6);font-size:13px;font-weight:600;margin-bottom:2px;">FROM</p>', unsafe_allow_html=True)
        cf_from = st.date_input("", format="DD/MM/YYYY", key="cf_from",
                                label_visibility="collapsed")
    with _cc4:
        st.markdown('<p style="color:rgba(255,255,255,0.6);font-size:13px;font-weight:600;margin-bottom:2px;">TO</p>', unsafe_allow_html=True)
        cf_to = st.date_input("", format="DD/MM/YYYY", key="cf_to",
                              label_visibility="collapsed")

    st.markdown('</div>', unsafe_allow_html=True)

    # Safe fallback — session state may be None on very first render before FY reset
    if cf_from is None:
        cf_from = _fy_s
    if cf_to is None:
        cf_to = _today

    # ── Expand / Collapse All ─────────────────────────────────────────────────
    # Match section toggles (_open suffix) and group toggles (_grp_ in key)
    # but NOT widget keys like cf_ent_radio / cf_from / cf_to.
    def _cf_toggle_keys():
        return [
            k for k in st.session_state
            if (k.startswith("cf_") or k.startswith("wcf_"))
            and ("_open" in k or "_grp_" in k)
        ]

    _ea_col, _ca_col = st.columns(2)
    with _ea_col:
        if st.button("⊞ Expand All", key="cf_expand_all",
                     use_container_width=True):
            for _k in _cf_toggle_keys():
                st.session_state[_k] = True
            st.rerun()
    with _ca_col:
        if st.button("⊟ Collapse All", key="cf_collapse_all",
                     use_container_width=True):
            for _k in _cf_toggle_keys():
                st.session_state[_k] = False
            st.rerun()

    # ── Data fetch ────────────────────────────────────────────────────────────
    def _get_cf(d_from, d_to):
        """Fetch cash flow; passes entity=None for All (triggers IC netting in fetch_cf)."""
        ds, de = str(d_from), str(d_to)
        ent = None if cf_ent == "All" else cf_ent
        return fetch_cf(ent, ds, de, DATABASE_FILE, financial_year=None)

    # ── Collapsible section helper ────────────────────────────────────────────
    def _cf_section(label, total, line_items, toggle_key, btn_key,
                    label_color="#1A1A2E", total_color="#1A1A2E", is_payout=False,
                    section_cls="", section_bg=""):
        """Renders one collapsible Cash Flow section."""
        h1, h2, h3 = st.columns([0.04, 0.71, 0.25])
        with h1:
            sign = "−" if st.session_state[toggle_key] else "+"
            if st.button(sign, key=btn_key):
                st.session_state[toggle_key] = not st.session_state[toggle_key]
                st.rerun()
        with h2:
            if section_cls:
                st.markdown(
                    f'<div class="cf-section-label {section_cls}">{label}</div>',
                    unsafe_allow_html=True)
            else:
                st.markdown(
                    f'<div style="font-weight:600; color:{label_color}; '
                    f'background:var(--bg-raised,#F4F6FA); padding:6px 8px; '
                    f'border-radius:6px;">{label}</div>', unsafe_allow_html=True)
        with h3:
            _tcls = "neg" if is_payout else "sky"
            _bg_style = f"background:{section_bg};" if section_bg else ""
            st.markdown(
                f'<div class="cf-section-total {_tcls}" style="{_bg_style}">{_inr(total)}</div>',
                unsafe_allow_html=True)
        if st.session_state[toggle_key]:
            for label_item, amount in line_items.items():
                _, r2, r3 = st.columns([0.04, 0.71, 0.25])
                _unc = label_item == "Uncategorized"
                _unc_cls = " unc" if _unc else ""
                with r2:
                    st.markdown(
                        f'<div class="cf-detail{_unc_cls}" style="background:#FFFFFF;color:#374151;">'
                        f'{label_item}</div>',
                        unsafe_allow_html=True)
                with r3:
                    st.markdown(
                        f'<div class="cf-dval{_unc_cls}" style="background:#FFFFFF;color:#374151;">'
                        f'{_inr(amount)}</div>',
                        unsafe_allow_html=True)
        st.markdown('<hr class="cf-divider-sm">', unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────────
    # MONTHLY VIEW
    # ─────────────────────────────────────────────────────────────────────────
    if cf_view == "Monthly":
        cf = _get_cf(cf_from, cf_to)

        # Uncategorized always last
        receipts_ordered = {k: v for k, v in cf["receipts"].items() if k != "Uncategorized"}
        if "Uncategorized" in cf["receipts"]:
            receipts_ordered["Uncategorized"] = cf["receipts"]["Uncategorized"]
        payouts_ordered = {k: v for k, v in cf["payouts"].items() if k != "Uncategorized"}
        if "Uncategorized" in cf["payouts"]:
            payouts_ordered["Uncategorized"] = cf["payouts"]["Uncategorized"]

        _cf_section(
            "Opening Balance", cf["total_opening"],
            {f"Bank ({k})": v for k, v in cf["opening_balances"].items()},
            "cf_ob_open", "btn_cf_ob", section_cls="cf-ob", section_bg="#DBEAFE"
        )
        # ── Add Receipts — GROUP → FINAL GROUP hierarchy ─────────────────
        _rh1, _rh2, _rh3 = st.columns([0.04, 0.71, 0.25])
        with _rh1:
            _rsign = "−" if st.session_state["cf_rec_open"] else "+"
            if st.button(_rsign, key="btn_cf_rec"):
                st.session_state["cf_rec_open"] = not st.session_state["cf_rec_open"]
                st.rerun()
        with _rh2:
            st.markdown('<div class="cf-section-label cf-rec">Add Receipts</div>',
                        unsafe_allow_html=True)
        with _rh3:
            st.markdown(
                f'<div class="cf-section-total pos" style="background:#EDE9FE;">'
                f'{_inr(cf["total_receipts"])}</div>',
                unsafe_allow_html=True)
        if st.session_state["cf_rec_open"]:
            for _grp_name, _fg_dict in sorted(
                cf.get("receipts_nested", {}).items(),
                key=lambda x: sum(x[1].values()), reverse=True
            ):
                _grp_total = sum(_fg_dict.values())
                _grp_key   = f"cf_grp_rec_{_grp_name}"
                if _grp_key not in st.session_state:
                    st.session_state[_grp_key] = False
                _g1, _g2, _g3 = st.columns([0.08, 0.67, 0.25])
                with _g1:
                    _gsign = "−" if st.session_state[_grp_key] else "+"
                    if st.button(_gsign, key=f"btn_{_grp_key}"):
                        st.session_state[_grp_key] = not st.session_state[_grp_key]
                        st.rerun()
                with _g2:
                    st.markdown(f'<div class="cf-grp-row rec">{_grp_name}</div>',
                                unsafe_allow_html=True)
                with _g3:
                    st.markdown(f'<div class="cf-grp-val rec">{_inr(_grp_total)}</div>',
                                unsafe_allow_html=True)
                if st.session_state[_grp_key]:
                    for _cat, _amt in sorted(_fg_dict.items(), key=lambda x: x[1], reverse=True):
                        _is_unc = _cat == "Uncategorized"
                        _sub_cls = " unc" if _is_unc else ""
                        _, _cr2, _cr3 = st.columns([0.08, 0.67, 0.25])
                        with _cr2:
                            st.markdown(f'<div class="cf-sub-row{_sub_cls}">{_cat}</div>',
                                        unsafe_allow_html=True)
                        with _cr3:
                            st.markdown(f'<div class="cf-sub-val{_sub_cls}">{_inr(_amt)}</div>',
                                        unsafe_allow_html=True)
        st.markdown('<hr class="cf-divider-sm">', unsafe_allow_html=True)

        # ── Less Payouts — GROUP → FINAL GROUP hierarchy ──────────────────
        _ph1, _ph2, _ph3 = st.columns([0.04, 0.71, 0.25])
        with _ph1:
            _psign = "−" if st.session_state["cf_pay_open"] else "+"
            if st.button(_psign, key="btn_cf_pay"):
                st.session_state["cf_pay_open"] = not st.session_state["cf_pay_open"]
                st.rerun()
        with _ph2:
            st.markdown('<div class="cf-section-label cf-pay">Less Payouts</div>',
                        unsafe_allow_html=True)
        with _ph3:
            st.markdown(
                f'<div class="cf-section-total neg" style="background:#FCE7F3;">'
                f'{_inr(cf["total_payouts"])}</div>',
                unsafe_allow_html=True)
        if st.session_state["cf_pay_open"]:
            for _grp_name, _fg_dict in sorted(
                cf.get("payouts_nested", {}).items(),
                key=lambda x: sum(x[1].values()), reverse=True
            ):
                _grp_total = sum(_fg_dict.values())
                _grp_key   = f"cf_grp_pay_{_grp_name}"
                if _grp_key not in st.session_state:
                    st.session_state[_grp_key] = False
                _g1, _g2, _g3 = st.columns([0.08, 0.67, 0.25])
                with _g1:
                    _gsign = "−" if st.session_state[_grp_key] else "+"
                    if st.button(_gsign, key=f"btn_{_grp_key}"):
                        st.session_state[_grp_key] = not st.session_state[_grp_key]
                        st.rerun()
                with _g2:
                    st.markdown(f'<div class="cf-grp-row pay">{_grp_name}</div>',
                                unsafe_allow_html=True)
                with _g3:
                    st.markdown(f'<div class="cf-grp-val pay">{_inr(_grp_total)}</div>',
                                unsafe_allow_html=True)
                if st.session_state[_grp_key]:
                    for _cat, _amt in sorted(_fg_dict.items(), key=lambda x: x[1], reverse=True):
                        _is_unc = _cat == "Uncategorized"
                        _sub_cls = " unc" if _is_unc else ""
                        _, _pr2, _pr3 = st.columns([0.08, 0.67, 0.25])
                        with _pr2:
                            st.markdown(f'<div class="cf-sub-row{_sub_cls}">{_cat}</div>',
                                        unsafe_allow_html=True)
                        with _pr3:
                            st.markdown(f'<div class="cf-sub-val{_sub_cls}">{_inr(_amt)}</div>',
                                        unsafe_allow_html=True)
        st.markdown('<hr class="cf-divider-sm">', unsafe_allow_html=True)

        # ── Reconciling lines: Interbank / Intercompany net ───────────────────
        for _recon_lbl, _recon_key in [
            ("Interbank (Net)",    "interbank_period_net"),
            ("Intercompany (Net)", "intercompany_period_net"),
        ]:
            _rval = cf.get(_recon_key) or 0
            if abs(_rval) > 1:
                _rcol = "#16A34A" if _rval >= 0 else "#DC2626"
                _rh1, _rh2, _rh3 = st.columns([0.04, 0.71, 0.25])
                with _rh2:
                    st.markdown(
                        f'<div style="padding:6px 12px; font-size:13px; font-weight:600; '
                        f'color:{_rcol}; background:#F8FAFF; border-radius:4px;">'
                        f'{_recon_lbl}</div>',
                        unsafe_allow_html=True)
                with _rh3:
                    st.markdown(
                        f'<div style="text-align:right; font-size:13px; color:{_rcol}; '
                        f'font-weight:600; padding:6px 0;">'
                        f'{fmt_cf(_rval)}</div>',
                        unsafe_allow_html=True)

        _cf_section(
            "Closing Balance", cf["total_closing"],
            {f"Bank ({k})": v for k, v in cf["closing_balances"].items()},
            "cf_cb_open", "btn_cf_cb", section_cls="cf-cb", section_bg="#DBEAFE"
        )

        # NET CASH FLOW — always visible
        _net     = cf["net_cash_flow"]
        _net_cls = "pos" if _net >= 0 else "neg"
        st.markdown(f"""
<div class="cf-net {_net_cls}">
  <span>NET CASH FLOW</span>
  <span>{_inr(_net)}</span>
</div>""", unsafe_allow_html=True)

        # TALLY CHECK — formula closing vs actual last DB balance
        # (closing is now computed, so compare against DB to detect data gaps)
        _tally     = cf["total_closing"]
        _db_actual = get_closing_balance(
            entity=cf_ent if cf_ent != "All" else None,
            date_from="1900-01-01",
            date_to=str(cf_to),
        )["total"] or 0
        _diff   = round(_tally - _db_actual, 2)
        _t_ok   = abs(_diff) <= 1
        _t_icon = "✅" if _t_ok else "⚠️"
        _t_msg  = "Tallied" if _t_ok else f"Diff: {_inr(abs(_diff))}"
        _t_cls  = "ok" if _t_ok else "err"
        st.markdown(f"""
<div class="cf-tally {_t_cls}">
  <strong>{_t_icon} Tally</strong>
  Computed: {_inr(_tally)} &nbsp;·&nbsp; DB: {_inr(_db_actual)}
  &nbsp;·&nbsp; <strong>{_t_msg}</strong>
</div>""", unsafe_allow_html=True)

        # Excel export
        _, _ec2, _ = st.columns([6, 1.5, 0.5])
        with _ec2:
            def _build_excel_monthly():
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                wb = Workbook()
                ws = wb.active
                ws.title = "Cash Flow"
                ws.column_dimensions["A"].width = 45
                ws.column_dimensions["B"].width = 22
                h_fill = PatternFill("solid", fgColor="F1F3F5")
                t_fill = PatternFill("solid", fgColor="E8F4FD")
                ws.append(["Particulars", "Amount (₹)"])
                for cell in ws[1]:
                    cell.font      = Font(bold=True, size=10, color="9CA3AF")
                    cell.fill      = h_fill
                    cell.alignment = Alignment(
                        horizontal="right" if cell.column > 1 else "left")

                def _xlr(label, val=None, style="normal", indent=0):
                    ws.append([("  " * indent) + label, val or None])
                    r = ws.max_row
                    if style == "header":
                        for c in ws[r]:
                            c.font = Font(bold=True, size=11)
                            c.fill = h_fill
                    elif style == "total":
                        for c in ws[r]:
                            c.font   = Font(bold=True, color="1E3A5F")
                            c.fill   = t_fill
                            c.border = Border(
                                top=Side(style="medium", color="4E8DF5"),
                                bottom=Side(style="medium", color="4E8DF5"))
                    elif style == "net":
                        pos = (val or 0) >= 0
                        for c in ws[r]:
                            c.font = Font(bold=True, size=12,
                                          color="2E7D32" if pos else "C62828")
                            c.fill = PatternFill("solid",
                                                 fgColor="F0FDF4" if pos else "FEF2F2")
                    if val is not None:
                        ws.cell(row=r, column=2).alignment = Alignment(
                            horizontal="right")

                _xlr("Opening Balance", style="header")
                for k, v in cf["opening_balances"].items():
                    _xlr(f"Bank ({k})", v, indent=1)
                _xlr("Total Opening Balance", cf["total_opening"], style="total")
                _xlr("Add Receipts:", style="header")
                for k, v in receipts_ordered.items():
                    _xlr(k, v, indent=1)
                _xlr("Total Receipts", cf["total_receipts"], style="total")
                _xlr("Less Payouts:", style="header")
                for k, v in payouts_ordered.items():
                    _xlr(k, v, indent=1)
                _xlr("Total Payouts", cf["total_payouts"], style="total")
                _xlr("Closing Balance", style="header")
                for k, v in cf["closing_balances"].items():
                    _xlr(f"Bank ({k})", v, indent=1)
                _xlr("Total Closing Balance", cf["total_closing"], style="total")
                _xlr("Net Cash Flow", cf["net_cash_flow"], style="net")
                out = io.BytesIO()
                wb.save(out)
                return out.getvalue()

            _xl_fname = (f"cashflow_{cf_ent}_{cf_from}_{cf_to}.xlsx"
                         .replace(" ", "_").replace("/", "-"))
            st.download_button(
                "⬇️ Export Excel", _build_excel_monthly(),
                file_name=_xl_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    # ─────────────────────────────────────────────────────────────────────────
    # WEEKLY VIEW
    # ─────────────────────────────────────────────────────────────────────────
    else:
        # W4 = current week Mon → today (inclusive, Live)
        # W1-W3 = three complete Mon-Sun weeks immediately before W4
        _days_since_mon = _today.weekday()  # Mon=0 … Sun=6
        _w4_start = _today - datetime.timedelta(days=_days_since_mon)
        _w4_end   = _today
        _w3_end   = _w4_start - datetime.timedelta(days=1)
        _w3_start = _w3_end   - datetime.timedelta(days=6)
        _w2_end   = _w3_start - datetime.timedelta(days=1)
        _w2_start = _w2_end   - datetime.timedelta(days=6)
        _w1_end   = _w2_start - datetime.timedelta(days=1)
        _w1_start = _w1_end   - datetime.timedelta(days=6)

        def _fmt_wk(d_from, d_to):
            f = d_from if isinstance(d_from, datetime.date) else datetime.date.fromisoformat(str(d_from))
            t = d_to   if isinstance(d_to,   datetime.date) else datetime.date.fromisoformat(str(d_to))
            return f"{f.strftime('%d %b')}–{t.strftime('%d %b')}"

        _w_ranges = [
            (_w1_start, _w1_end),
            (_w2_start, _w2_end),
            (_w3_start, _w3_end),
            (_w4_start, _w4_end),
        ]
        _w_labels = [
            f"W1 {_fmt_wk(_w1_start, _w1_end)}",
            f"W2 {_fmt_wk(_w2_start, _w2_end)}",
            f"W3 {_fmt_wk(_w3_start, _w3_end)}",
            f"W4 {_fmt_wk(_w4_start, _w4_end)}",
        ]
        _w_data = [_get_cf(w[0], w[1]) for w in _w_ranges]

        # Rolling balance with DB validation
        _ro = [0.0] * 4
        _rc = [0.0] * 4
        _ro[0] = _w_data[0]["total_opening"]
        for _wi in range(4):
            _computed = _ro[_wi] + _w_data[_wi]["total_receipts"] - _w_data[_wi]["total_payouts"]
            _rc[_wi]  = _computed
            _db_close = sum(_w_data[_wi]["closing_balances"].values())
            _w_data[_wi]["computed_closing"] = _computed
            _w_data[_wi]["db_closing"]       = _db_close
            _w_data[_wi]["tally_diff"]       = round(_computed - _db_close, 2)
            if _wi < 3:
                _ro[_wi + 1] = _computed

        # Per-bank opening/closing — use actual DB balances (last balance on or before week end).
        # W(n+1) opening = W(n) actual DB closing.  No proportional approximation.
        _bank_close = {_wi: dict(_w_data[_wi]["closing_balances"]) for _wi in range(4)}
        _bank_open  = {0: dict(_w_data[0]["opening_balances"])}
        for _wi in range(3):
            _bank_open[_wi + 1] = dict(_bank_close[_wi])

        # DB-sourced totals — used for ALL header rows so they always match the detail sums.
        _ob_totals = [sum(_bank_open[_wi].values())  for _wi in range(4)]
        _cb_totals = [sum(_bank_close[_wi].values()) for _wi in range(4)]

        # Tally diffs: formula-computed closing vs actual last DB balance per week
        _db_actual_close = [
            get_closing_balance(
                entity=cf_ent if cf_ent != "All" else None,
                date_from="1900-01-01",
                date_to=str(_w_ranges[_wi][1]),
            )["total"] or 0
            for _wi in range(4)
        ]
        for _wi in range(4):
            _w_data[_wi]["tally_diff"] = round(_cb_totals[_wi] - _db_actual_close[_wi], 2)

        # All unique categories across all weeks — Uncategorized always last
        _all_rec_cats = []
        for _wd in _w_data:
            for _cat in _wd["receipts"]:
                if _cat not in _all_rec_cats:
                    _all_rec_cats.append(_cat)
        _all_rec_cats = [c for c in _all_rec_cats if c != "Uncategorized"]
        if any("Uncategorized" in _wd["receipts"] for _wd in _w_data):
            _all_rec_cats.append("Uncategorized")

        _all_pay_cats = []
        for _wd in _w_data:
            for _cat in _wd["payouts"]:
                if _cat not in _all_pay_cats:
                    _all_pay_cats.append(_cat)
        _all_pay_cats = [c for c in _all_pay_cats if c != "Uncategorized"]
        if any("Uncategorized" in _wd["payouts"] for _wd in _w_data):
            _all_pay_cats.append("Uncategorized")

        # Single column layout used for every row in the weekly table
        CF_COLS = [0.04, 0.32, 0.12, 0.12, 0.12, 0.12, 0.12]
        _wk_cls = ["std", "std", "std", "std"]  # uniform week column headers

        # ── Colour palette ────────────────────────────────────────────────
        _WK_BG  = ["#F0F7FF", "#FFFFFF", "#F0F7FF", "#FFFFFF"]  # W1-W4 alternating
        _TOT_BG = "#EFF6FF"
        _SEC = {
            "opening":  {"bg": "#DBEAFE", "text": "#1E40AF"},
            "receipts": {"bg": "#EDE9FE", "text": "#5B21B6"},
            "payouts":  {"bg": "#FCE7F3", "text": "#9D174D"},
            "closing":  {"bg": "#DBEAFE", "text": "#1E40AF"},
            "net_pos":  {"bg": "#F0FDF4", "text": "#14532D"},
            "net_neg":  {"bg": "#FFF1F2", "text": "#9D174D"},
        }

        # ── Column header row ─────────────────────────────────────────────
        _hdr = st.columns(CF_COLS)
        with _hdr[1]:
            st.markdown(
                '<div style="font-size:13px;font-weight:700;letter-spacing:0.1em;'
                'text-transform:uppercase;color:var(--tx-4,#96A3B4);padding:6px 0;">'
                'Particulars</div>', unsafe_allow_html=True)
        for _i, _lbl in enumerate(_w_labels):
            with _hdr[_i + 2]:
                st.markdown(
                    f'<div class="cf-wk-hdr {_wk_cls[_i]}">{_lbl}</div>',
                    unsafe_allow_html=True)
        with _hdr[6]:
            st.markdown(
                '<div class="cf-wk-hdr ttl">TOTAL</div>',
                unsafe_allow_html=True)
        st.markdown(
            '<hr style="margin:2px 0;border:none;border-top:2px solid var(--blue-400,#3B82F6);">',
            unsafe_allow_html=True)

        # ── Opening Balance ───────────────────────────────────────────────
        _ob_r = st.columns(CF_COLS)
        with _ob_r[0]:
            _sign = "−" if st.session_state["wcf_ob_open"] else "+"
            if st.button(_sign, key="wbtn_ob"):
                st.session_state["wcf_ob_open"] = not st.session_state["wcf_ob_open"]
                st.rerun()
        with _ob_r[1]:
            st.markdown(
                '<div class="cf-section-label cf-ob">Opening Balance</div>',
                unsafe_allow_html=True)
        for _i in range(4):
            with _ob_r[_i + 2]:
                st.markdown(
                    f'<div class="cf-wk-val bold" style="background:{_SEC["opening"]["bg"]};'
                    f'color:{_SEC["opening"]["text"]};">{_inr(_ob_totals[_i])}</div>',
                    unsafe_allow_html=True)
        with _ob_r[6]:
            st.markdown(
                f'<div class="cf-wk-val bold" style="background:{_SEC["opening"]["bg"]};'
                f'color:{_SEC["opening"]["text"]};">{_inr(_ob_totals[0])}</div>',
                unsafe_allow_html=True)
        if st.session_state["wcf_ob_open"]:
            for _bk in sorted(_bank_open[0]):
                _br = st.columns(CF_COLS)
                with _br[1]:
                    st.markdown(f'<div class="cf-detail">Bank ({_bk})</div>',
                                unsafe_allow_html=True)
                for _i in range(4):
                    _bov = _bank_open[_i].get(_bk, 0)
                    with _br[_i + 2]:
                        _dash_cls = " dim" if not _bov else ""
                        st.markdown(
                            f'<div class="cf-wk-val{_dash_cls}">'
                            f'{_inr(_bov) if _bov else "—"}</div>',
                            unsafe_allow_html=True)
                with _br[6]:
                    st.markdown(
                        f'<div class="cf-wk-val total">'
                        f'{_inr(_bank_open[0].get(_bk, 0))}</div>',
                        unsafe_allow_html=True)
        st.markdown(
            '<hr class="cf-divider-sm">',
            unsafe_allow_html=True)

        # ── Add Receipts ──────────────────────────────────────────────────
        _rec_r = st.columns(CF_COLS)
        with _rec_r[0]:
            _sign = "−" if st.session_state["wcf_rec_open"] else "+"
            if st.button(_sign, key="wbtn_rec"):
                st.session_state["wcf_rec_open"] = not st.session_state["wcf_rec_open"]
                st.rerun()
        with _rec_r[1]:
            st.markdown(
                '<div class="cf-section-label cf-rec">Add Receipts</div>',
                unsafe_allow_html=True)
        for _i, _wd in enumerate(_w_data):
            with _rec_r[_i + 2]:
                st.markdown(
                    f'<div class="cf-wk-val bold" style="background:{_SEC["receipts"]["bg"]};'
                    f'color:{_SEC["receipts"]["text"]};">{_inr(_wd["total_receipts"])}</div>',
                    unsafe_allow_html=True)
        with _rec_r[6]:
            st.markdown(
                f'<div class="cf-wk-val bold" style="background:{_SEC["receipts"]["bg"]};'
                f'color:{_SEC["receipts"]["text"]};">'
                f'{_inr(sum(_wd["total_receipts"] for _wd in _w_data))}</div>',
                unsafe_allow_html=True)
        if st.session_state["wcf_rec_open"]:
            _wk_rec_nested = [_wd.get("receipts_nested", {}) for _wd in _w_data]
            _all_rec_groups = sorted(
                set(g for _wn in _wk_rec_nested for g in _wn),
                key=lambda g: sum(sum(_wn.get(g, {}).values()) for _wn in _wk_rec_nested),
                reverse=True,
            )
            for _grp in _all_rec_groups:
                _grp_key = f"wcf_grp_rec_{_grp}"
                if _grp_key not in st.session_state:
                    st.session_state[_grp_key] = False
                _wgt = [sum(_wn.get(_grp, {}).values()) for _wn in _wk_rec_nested]
                _wgt_total = sum(_wgt)
                _gr = st.columns(CF_COLS)
                with _gr[0]:
                    _gsign = "−" if st.session_state[_grp_key] else "+"
                    if st.button(_gsign, key=f"btn_{_grp_key}"):
                        st.session_state[_grp_key] = not st.session_state[_grp_key]
                        st.rerun()
                with _gr[1]:
                    st.markdown(
                        f'<div class="cf-wk-grp-lbl rec" style="padding-left:14px;">{_grp}</div>',
                        unsafe_allow_html=True)
                for _i in range(4):
                    with _gr[_i + 2]:
                        _gv = _wgt[_i]
                        st.markdown(
                            f'<div class="cf-wk-grp-val rec">{_inr(_gv) if _gv else "—"}</div>',
                            unsafe_allow_html=True)
                with _gr[6]:
                    st.markdown(f'<div class="cf-wk-grp-val rec total">{_inr(_wgt_total)}</div>',
                                unsafe_allow_html=True)
                if st.session_state[_grp_key]:
                    _cats_in_grp = sorted(
                        set(cat for _wn in _wk_rec_nested for cat in _wn.get(_grp, {})),
                        key=lambda c: sum(_wn.get(_grp, {}).get(c, 0) for _wn in _wk_rec_nested),
                        reverse=True,
                    )
                    for _cat in _cats_in_grp:
                        _is_unc = _cat == "Uncategorized"
                        _sub_cls = " unc" if _is_unc else ""
                        _cc = st.columns(CF_COLS)
                        with _cc[1]:
                            st.markdown(
                                f'<div class="cf-wk-sub-lbl{_sub_cls}" style="padding-left:32px;">'
                                f'{_cat}</div>',
                                unsafe_allow_html=True)
                        _cat_total = 0
                        for _i in range(4):
                            _amt = _wk_rec_nested[_i].get(_grp, {}).get(_cat, 0)
                            _cat_total += _amt
                            _cell_bg = _WK_BG[_i]
                            _cell_st = f"background:{_cell_bg};color:#374151;"
                            with _cc[_i + 2]:
                                st.markdown(
                                    f'<div class="cf-wk-sub-val{_sub_cls}" style="{_cell_st}">'
                                    f'{_inr(_amt) if _amt else "—"}</div>',
                                    unsafe_allow_html=True)
                        with _cc[6]:
                            st.markdown(
                                f'<div class="cf-wk-sub-val total" style="background:{_TOT_BG};">'
                                f'{_inr(_cat_total)}</div>',
                                unsafe_allow_html=True)
        st.markdown(
            '<hr class="cf-divider-sm">',
            unsafe_allow_html=True)

        # ── Less Payouts ──────────────────────────────────────────────────
        _pay_r = st.columns(CF_COLS)
        with _pay_r[0]:
            _sign = "−" if st.session_state["wcf_pay_open"] else "+"
            if st.button(_sign, key="wbtn_pay"):
                st.session_state["wcf_pay_open"] = not st.session_state["wcf_pay_open"]
                st.rerun()
        with _pay_r[1]:
            st.markdown(
                '<div class="cf-section-label cf-pay">Less Payouts</div>',
                unsafe_allow_html=True)
        for _i, _wd in enumerate(_w_data):
            with _pay_r[_i + 2]:
                st.markdown(
                    f'<div class="cf-wk-val bold" style="background:{_SEC["payouts"]["bg"]};'
                    f'color:{_SEC["payouts"]["text"]};">{_inr(_wd["total_payouts"])}</div>',
                    unsafe_allow_html=True)
        with _pay_r[6]:
            st.markdown(
                f'<div class="cf-wk-val bold" style="background:{_SEC["payouts"]["bg"]};'
                f'color:{_SEC["payouts"]["text"]};">'
                f'{_inr(sum(_wd["total_payouts"] for _wd in _w_data))}</div>',
                unsafe_allow_html=True)
        if st.session_state["wcf_pay_open"]:
            _wk_pay_nested = [_wd.get("payouts_nested", {}) for _wd in _w_data]
            _all_pay_groups = sorted(
                set(g for _wn in _wk_pay_nested for g in _wn),
                key=lambda g: sum(sum(_wn.get(g, {}).values()) for _wn in _wk_pay_nested),
                reverse=True,
            )
            for _grp in _all_pay_groups:
                _grp_key = f"wcf_grp_pay_{_grp}"
                if _grp_key not in st.session_state:
                    st.session_state[_grp_key] = False
                _wpt = [sum(_wn.get(_grp, {}).values()) for _wn in _wk_pay_nested]
                _wpt_total = sum(_wpt)
                _gr = st.columns(CF_COLS)
                with _gr[0]:
                    _gsign = "−" if st.session_state[_grp_key] else "+"
                    if st.button(_gsign, key=f"btn_{_grp_key}"):
                        st.session_state[_grp_key] = not st.session_state[_grp_key]
                        st.rerun()
                with _gr[1]:
                    st.markdown(
                        f'<div class="cf-wk-grp-lbl pay" style="padding-left:14px;">{_grp}</div>',
                        unsafe_allow_html=True)
                for _i in range(4):
                    with _gr[_i + 2]:
                        _pv = _wpt[_i]
                        st.markdown(
                            f'<div class="cf-wk-grp-val pay">{_inr(_pv) if _pv else "—"}</div>',
                            unsafe_allow_html=True)
                with _gr[6]:
                    st.markdown(f'<div class="cf-wk-grp-val pay total">{_inr(_wpt_total)}</div>',
                                unsafe_allow_html=True)
                if st.session_state[_grp_key]:
                    _cats_in_grp = sorted(
                        set(cat for _wn in _wk_pay_nested for cat in _wn.get(_grp, {})),
                        key=lambda c: sum(_wn.get(_grp, {}).get(c, 0) for _wn in _wk_pay_nested),
                        reverse=True,
                    )
                    for _cat in _cats_in_grp:
                        _is_unc = _cat == "Uncategorized"
                        _sub_cls = " unc" if _is_unc else ""
                        _cc = st.columns(CF_COLS)
                        with _cc[1]:
                            st.markdown(
                                f'<div class="cf-wk-sub-lbl{_sub_cls}" style="padding-left:32px;">'
                                f'{_cat}</div>',
                                unsafe_allow_html=True)
                        _cat_total = 0
                        for _i in range(4):
                            _amt = _wk_pay_nested[_i].get(_grp, {}).get(_cat, 0)
                            _cat_total += _amt
                            _cell_bg = _WK_BG[_i]
                            _cell_st = f"background:{_cell_bg};color:#374151;"
                            with _cc[_i + 2]:
                                st.markdown(
                                    f'<div class="cf-wk-sub-val{_sub_cls}" style="{_cell_st}">'
                                    f'{_inr(_amt) if _amt else "—"}</div>',
                                    unsafe_allow_html=True)
                        with _cc[6]:
                            st.markdown(
                                f'<div class="cf-wk-sub-val ptotal" style="background:{_TOT_BG};">'
                                f'{_inr(_cat_total)}</div>',
                                unsafe_allow_html=True)
        st.markdown(
            '<hr class="cf-divider-sm">',
            unsafe_allow_html=True)

        # ── Reconciling lines: Interbank / Intercompany net per week ─────
        _WK_RECON_BG = ["#F0F7FF", "#FFFFFF", "#F0F7FF", "#FFFFFF"]
        for _recon_lbl, _recon_key in [
            ("Interbank (Net)",    "interbank_period_net"),
            ("Intercompany (Net)", "intercompany_period_net"),
        ]:
            _week_vals   = [_wd.get(_recon_key) or 0 for _wd in _w_data]
            _grand_total = sum(_week_vals)
            if any(abs(_v) > 1 for _v in _week_vals):
                _rr = st.columns(CF_COLS)
                with _rr[1]:
                    _rc = "#16A34A" if _grand_total >= 0 else "#DC2626"
                    st.markdown(
                        f'<div style="padding:6px 12px; font-size:13px; font-weight:600; '
                        f'color:{_rc}; background:#F8FAFF; border-radius:4px;">'
                        f'{_recon_lbl}</div>',
                        unsafe_allow_html=True)
                for _i, _v in enumerate(_week_vals):
                    with _rr[_i + 2]:
                        _vc = "#16A34A" if _v >= 0 else "#DC2626"
                        _vd = fmt_cf(_v) if abs(_v) > 1 else "—"
                        st.markdown(
                            f'<div style="text-align:right; font-size:12px; color:{_vc}; '
                            f'font-weight:600; background:{_WK_RECON_BG[_i]}; padding:6px;">'
                            f'{_vd}</div>',
                            unsafe_allow_html=True)
                with _rr[6]:
                    _tc = "#16A34A" if _grand_total >= 0 else "#DC2626"
                    st.markdown(
                        f'<div style="text-align:right; font-size:12px; color:{_tc}; '
                        f'font-weight:700; background:#F0F4FF; padding:6px;">'
                        f'{fmt_cf(_grand_total)}</div>',
                        unsafe_allow_html=True)

        # ── Closing Balance ───────────────────────────────────────────────
        _cb_r = st.columns(CF_COLS)
        with _cb_r[0]:
            _sign = "−" if st.session_state["wcf_cb_open"] else "+"
            if st.button(_sign, key="wbtn_cb"):
                st.session_state["wcf_cb_open"] = not st.session_state["wcf_cb_open"]
                st.rerun()
        with _cb_r[1]:
            st.markdown(
                '<div class="cf-section-label cf-cb">Closing Balance</div>',
                unsafe_allow_html=True)
        for _i in range(4):
            with _cb_r[_i + 2]:
                st.markdown(
                    f'<div class="cf-wk-val bold" style="background:{_SEC["closing"]["bg"]};'
                    f'color:{_SEC["closing"]["text"]};">{_inr(_cb_totals[_i])}</div>',
                    unsafe_allow_html=True)
        with _cb_r[6]:
            st.markdown(
                f'<div class="cf-wk-val bold" style="background:{_SEC["closing"]["bg"]};'
                f'color:{_SEC["closing"]["text"]};">{_inr(_cb_totals[3])}</div>',
                unsafe_allow_html=True)
        if st.session_state["wcf_cb_open"]:
            for _bk in sorted(_bank_close[0]):
                _cbr = st.columns(CF_COLS)
                with _cbr[1]:
                    st.markdown(f'<div class="cf-detail">Bank ({_bk})</div>',
                                unsafe_allow_html=True)
                for _i in range(4):
                    _bcv = _bank_close[_i].get(_bk, 0)
                    with _cbr[_i + 2]:
                        _d2 = " dim" if not _bcv else ""
                        st.markdown(
                            f'<div class="cf-wk-val{_d2}">'
                            f'{_inr(_bcv) if _bcv else "—"}</div>',
                            unsafe_allow_html=True)
                with _cbr[6]:
                    st.markdown(
                        f'<div class="cf-wk-val bold total">'
                        f'{_inr(_bank_close[3].get(_bk, 0))}</div>',
                        unsafe_allow_html=True)
        st.markdown(
            '<hr class="cf-divider-sm">',
            unsafe_allow_html=True)

        # ── Net Cash Flow per week — always visible ───────────────────────
        _wnet_r = st.columns(CF_COLS)
        with _wnet_r[1]:
            st.markdown(
                '<div class="cf-wk-net-label">NET CASH FLOW</div>',
                unsafe_allow_html=True)
        _total_wnet = sum(
            _w_data[_wi]["total_receipts"] - _w_data[_wi]["total_payouts"]
            for _wi in range(4))
        for _i, _wd in enumerate(_w_data):
            _wn     = _wd["total_receipts"] - _wd["total_payouts"]
            _wn_sec = "net_pos" if _wn >= 0 else "net_neg"
            with _wnet_r[_i + 2]:
                st.markdown(
                    f'<div class="cf-wk-val bold" style="background:{_SEC[_wn_sec]["bg"]};'
                    f'color:{_SEC[_wn_sec]["text"]};font-weight:700;">'
                    f'{_inr(_wn)}</div>', unsafe_allow_html=True)
        _tn_sec = "net_pos" if _total_wnet >= 0 else "net_neg"
        with _wnet_r[6]:
            st.markdown(
                f'<div class="cf-wk-val bold" style="background:{_SEC[_tn_sec]["bg"]};'
                f'color:{_SEC[_tn_sec]["text"]};font-weight:700;">'
                f'{_inr(_total_wnet)}</div>', unsafe_allow_html=True)
        st.markdown(
            '<hr style="margin:2px 0;border:none;border-top:3px solid var(--blue-400,#3B82F6);">',
            unsafe_allow_html=True)

        # ── Tally per week — computed vs DB ──────────────────────────────
        _tally_r = st.columns(CF_COLS)
        with _tally_r[1]:
            st.markdown(
                '<div class="cf-wk-tally-label">Tally (Computed vs DB)</div>',
                unsafe_allow_html=True)
        for _wi in range(4):
            _tdif = _w_data[_wi]["tally_diff"]
            _tok  = abs(_tdif) <= 1
            _tic  = "✅" if _tok else "⚠️"
            _tip  = "Tallied" if _tok else fmt_cf(abs(_tdif))
            _tbg_t = "var(--green-50,#F0FDF4)" if _tok else "var(--amber-50,#FFFBEB)"
            with _tally_r[_wi + 2]:
                st.markdown(
                    f'<div style="text-align:center;padding:3px 2px;background:{_tbg_t};'
                    f'border-radius:4px;font-size:13px;" title="{_tip}">{_tic}</div>',
                    unsafe_allow_html=True)
        _total_tdif = round(_cb_totals[3] - _db_actual_close[3], 2)
        with _tally_r[6]:
            _gt_ok  = abs(_total_tdif) <= 1
            _gt_tip = "Tallied" if _gt_ok else fmt_cf(abs(_total_tdif))
            _gt_bg  = "var(--green-50,#F0FDF4)" if _gt_ok else "var(--amber-50,#FFFBEB)"
            st.markdown(
                f'<div style="text-align:center;padding:3px 2px;background:{_gt_bg};'
                f'border-radius:4px;font-size:13px;" title="{_gt_tip}">'
                f'{"✅" if _gt_ok else "⚠️"}</div>',
                unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# INVESTMENTS
# ══════════════════════════════════════════════════════════════════════════════
elif selected_tab == "Investments":

    st.markdown("## Investment Portfolio")

    # ── Filters ──────────────────────────────────────────────────────────────
    f1, f2, f3, f4 = st.columns([2, 2, 2, 2])
    with f1:
        inv_entity = st.radio(
            "Entity", ["All", "Stores", "Ventures"],
            horizontal=True, key="inv_entity")
    with f2:
        inv_from = st.date_input(
            "From", value=_first_of_month, key="inv_from")
    with f3:
        inv_to = st.date_input(
            "To", value=_today, key="inv_to")
    with f4:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Refresh", key="inv_refresh"):
            st.rerun()

    # ── Get data ──────────────────────────────────────────────────────────────
    entity_filter = inv_entity if inv_entity != "All" else None
    inv_data = get_investment_summary(
        entity=entity_filter,
        date_from=str(inv_from),
        date_to=str(inv_to),
    )

    fds = [r for r in inv_data if r["scheme_type"] == "FD"]
    mfs = [r for r in inv_data if r["scheme_type"] == "MF"]

    # ── Import from Excel ─────────────────────────────────────────────────────
    with st.expander("📥 Import Investments from Excel", expanded=False):
        _dl_col, _up_col = st.columns([1, 2])
        with _dl_col:
            import io as _io
            import pandas as _pde
            _tmpl_df = _pde.DataFrame(columns=[
                "Date", "Scheme Name", "Scheme Number",
                "Type", "Amount", "Entry Type", "Entity",
            ])
            _tmpl_buf = _io.BytesIO()
            _tmpl_df.to_excel(_tmpl_buf, index=False)
            _tmpl_buf.seek(0)
            st.download_button(
                "⬇ Download Template",
                data=_tmpl_buf,
                file_name="investment_import_template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with _up_col:
            _inv_upload = st.file_uploader(
                "Upload filled Excel",
                type=["xlsx", "xls"],
                key="inv_excel_upload",
            )
            if _inv_upload is not None:
                if st.button("📤 Import", key="inv_do_import", use_container_width=True):
                    import tempfile as _tf, os as _osi
                    with _tf.NamedTemporaryFile(delete=False, suffix=".xlsx") as _tmp:
                        _tmp.write(_inv_upload.getvalue())
                        _tmp_path = _tmp.name
                    _result = import_investment_excel(_tmp_path)
                    try:
                        _osi.unlink(_tmp_path)
                    except Exception:
                        pass
                    if _result["errors"] == 0 and _result["skipped"] == 0:
                        st.success(f"Imported {_result['inserted']} rows successfully.")
                    else:
                        st.warning(
                            f"Inserted: {_result['inserted']} · "
                            f"Skipped: {_result['skipped']} · "
                            f"Errors: {_result['errors']}"
                        )
                        for _err in _result["error_rows"][:10]:
                            st.caption(_err)
                    if _result["inserted"] > 0:
                        st.cache_data.clear()
                        try:
                            st.rerun()
                        except AttributeError:
                            st.experimental_rerun()

    # ── Summary KPI strip ─────────────────────────────────────────────────────
    _inv_kpis   = get_investment_kpis(
        entity=inv_entity if inv_entity != "All" else None
    )
    _ik_opening = _inv_kpis["opening"]
    _ik_mf      = _inv_kpis["mf_balance"]
    _ik_fd      = _inv_kpis["fd_balance"]
    _ik_closing = _inv_kpis["closing"]
    _ik_cls     = "#16A34A" if _ik_closing >= _ik_opening else "#DC2626"

    k1, k2, k3, k4 = st.columns(4)
    for _col, _label, _value, _color in [
        (k1, "OPENING BALANCE", _ik_opening, "#1B2B4B"),
        (k2, "MUTUAL FUND",     _ik_mf,      "#5B21B6"),
        (k3, "FIXED DEPOSIT",   _ik_fd,      "#1E40AF"),
        (k4, "CLOSING BALANCE", _ik_closing, _ik_cls),
    ]:
        with _col:
            st.markdown(f"""
<div style="background:#FFFFFF;border-radius:10px;padding:16px 20px;
     border:1px solid #E8ECF0;border-top:3px solid {_color};
     box-shadow:0 1px 3px rgba(0,0,0,0.04);">
  <div style="font-size:10px;font-weight:700;color:#8896A5;letter-spacing:0.1em;
       text-transform:uppercase;margin-bottom:6px;">{_label}</div>
  <div style="font-size:20px;font-weight:700;color:{_color};">{fmt_cr(_value)}</div>
</div>
""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Cash at Stores ────────────────────────────────────────────────────────
    st.markdown('<div class="section-header" style="margin-bottom:10px;">Cash at Stores</div>',
                unsafe_allow_html=True)
    _cas_col1, _cas_col2, _cas_col3 = st.columns([2, 2, 3])
    with _cas_col1:
        _cur_cas = get_cash_at_stores()
        st.markdown(f"""
<div style="background:#FFFFFF;border-radius:10px;padding:16px 20px;
     border:1px solid #E8ECF0;border-top:3px solid #0EA5E9;
     box-shadow:0 1px 3px rgba(0,0,0,0.04);">
  <div style="font-size:10px;font-weight:700;color:#8896A5;letter-spacing:0.1em;
       text-transform:uppercase;margin-bottom:6px;">Current Value</div>
  <div style="font-size:22px;font-weight:700;color:#0EA5E9;">₹{_cur_cas:,.0f}</div>
</div>
""", unsafe_allow_html=True)
    with _cas_col2:
        _new_cas = st.number_input(
            "Update Cash at Stores (₹)", min_value=0.0, step=1000.0,
            value=float(get_cash_at_stores()), format="%.0f",
            key="cas_input"
        )
        if st.button("Save", key="cas_save"):
            set_cash_at_stores(_new_cas)
            st.success(f"Saved ₹{_new_cas:,.0f}")
            st.rerun()
    st.markdown("<br>", unsafe_allow_html=True)

    # ── B4: Investment Portfolio — movements from actual transactions ─────────
    st.markdown('<div class="section-header" style="margin-bottom:10px;">Investment Portfolio</div>',
                unsafe_allow_html=True)

    inv_movements  = get_investment_movements(
        entity=entity_filter,
        date_from=str(inv_from),
        date_to=str(inv_to),
    )
    inv_register   = get_investment_register()
    manual_totals  = get_manual_investment_totals()
    register_map   = {r["scheme_name"]: r for r in inv_register}

    _rows_display = []
    _seen_schemes = set()

    # Rows from bank transactions tagged with main_group=INVESTMENT
    for _mov in inv_movements:
        _sn = _mov["scheme_name"]
        _seen_schemes.add(_sn)
        _reg     = register_map.get(_sn, {})
        _opening = _reg.get("opening_value", 0) or 0
        _current = _opening + (_mov["invested"] or 0) - (_mov["redeemed"] or 0)
        _rows_display.append({
            "Scheme":        _sn,
            "Number":        _mov["scheme_number"],
            "Type":          _mov["scheme_type"],
            "Entity":        _mov["entity"],
            "Source":        "Bank Txn",
            "Opening (₹)":  fmt_inr(_opening),
            "Invested (₹)": fmt_inr(_mov["invested"]) if _mov["invested"] else "—",
            "Redeemed (₹)": fmt_inr(_mov["redeemed"]) if _mov["redeemed"] else "—",
            "Current (₹)":  fmt_inr(_current),
            "Txns":          _mov["txn_count"],
        })

    # Rows from manual entries (off-statement FD/MF)
    for _mt in manual_totals:
        _sn = _mt["scheme_name"]
        _seen_schemes.add(_sn)
        _reg      = register_map.get(_sn, {})
        _opening  = _reg.get("opening_value", 0) or 0
        _invested = _mt["invested"]  or 0
        _redeemed = _mt["redeemed"]  or 0
        _current  = _opening + _invested - _redeemed
        _rows_display.append({
            "Scheme":        _sn,
            "Number":        "—",
            "Type":          _mt["scheme_type"],
            "Entity":        _mt["entity"],
            "Source":        "Manual",
            "Opening (₹)":  fmt_inr(_opening),
            "Invested (₹)": fmt_inr(_invested) if _invested else "—",
            "Redeemed (₹)": fmt_inr(_redeemed) if _redeemed else "—",
            "Current (₹)":  fmt_inr(_current),
            "Txns":          _mt["entry_count"],
        })

    # Static register entries with no transactions and no manual entries
    for _r in inv_register:
        if _r["scheme_name"] not in _seen_schemes:
            _rows_display.append({
                "Scheme":        _r["scheme_name"],
                "Number":        "—",
                "Type":          _r["scheme_type"],
                "Entity":        _r["entity"],
                "Source":        "Register",
                "Opening (₹)":  fmt_inr(_r["opening_value"] or 0),
                "Invested (₹)": "—",
                "Redeemed (₹)": "—",
                "Current (₹)":  fmt_inr(_r["opening_value"] or 0),
                "Txns":          0,
            })

    if _rows_display:
        _df_inv_portfolio = pd.DataFrame(_rows_display)
        st.dataframe(_df_inv_portfolio, use_container_width=True, hide_index=True)

        # Total row
        _total_current = 0.0
        for _row in _rows_display:
            _cv = _row.get("Current (₹)", "")
            if _cv and _cv not in ("—", ""):
                try:
                    _total_current += float(
                        str(_cv).replace("₹", "").replace(",", "").strip()
                    )
                except Exception:
                    pass
        st.markdown(f"""
<div style="background:#1B2B4B;color:#FFFFFF;border-radius:8px;
     padding:12px 20px;display:flex;justify-content:space-between;
     margin-top:8px;">
    <b>TOTAL INVESTMENT VALUE</b>
    <b>{fmt_inr(_total_current)}</b>
</div>""", unsafe_allow_html=True)
    else:
        st.info("No investment data found.")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Manage Schemes & Opening Values ──────────────────────────────────────
    with st.expander("⚙️ Manage Schemes & Opening Values", expanded=False):
        st.caption(
            "Set the opening value for each scheme as on 1st April. "
            "These are saved permanently in the database.")

        all_schemes = get_investment_register()

        st.markdown("##### Update Opening Values")
        for _scheme in all_schemes:
            _ec1, _ec2, _ec3, _ec4 = st.columns([2, 2, 2, 1])
            with _ec1:
                st.markdown(
                    f'<div style="padding-top:8px;font-weight:600;">'
                    f'{_scheme["scheme_name"]} '
                    f'<span style="color:#8896A5;font-size:11px;">'
                    f'({_scheme["scheme_type"]} · {_scheme["entity"]})'
                    f'</span></div>', unsafe_allow_html=True)
            with _ec2:
                _new_val = st.number_input(
                    "Opening Value (₹)",
                    value=float(_scheme["opening_value"] or 0),
                    min_value=0.0,
                    step=10000.0,
                    key=f"ov_{_scheme['scheme_name']}",
                    label_visibility="collapsed")
            with _ec3:
                _new_ent = st.selectbox(
                    "Entity",
                    ["Stores", "Ventures"],
                    index=0 if _scheme["entity"] == "Stores" else 1,
                    key=f"ent_{_scheme['scheme_name']}",
                    label_visibility="collapsed")
            with _ec4:
                if st.button("💾 Save", key=f"save_{_scheme['scheme_name']}"):
                    update_investment_opening(
                        _scheme["scheme_name"], _new_val,
                        "2026-04-01", _new_ent)
                    st.success(f"Saved {_scheme['scheme_name']}")
                    st.rerun()

        st.markdown("---")
        st.markdown("##### Add New Scheme")
        _na1, _na2, _na3, _na4, _na5 = st.columns([2, 1, 1, 2, 1])
        with _na1:
            _new_name = st.text_input(
                "Scheme Name", placeholder="e.g. FD-5 or MF-Scheme-4",
                key="new_scheme_name")
        with _na2:
            _new_type = st.selectbox("Type", ["FD", "MF"], key="new_scheme_type")
        with _na3:
            _new_ent2 = st.selectbox(
                "Entity", ["Stores", "Ventures"], key="new_scheme_entity")
        with _na4:
            _new_opening = st.number_input(
                "Opening Value (₹)", min_value=0.0,
                step=10000.0, key="new_scheme_opening")
        with _na5:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("➕ Add", key="add_scheme_btn"):
                if _new_name.strip():
                    add_investment_scheme(
                        _new_name.strip(), _new_type,
                        _new_ent2, _new_opening, "2026-04-01")
                    st.success(f"Added {_new_name}")
                    st.rerun()
                else:
                    st.warning("Enter a scheme name first.")

        st.markdown("---")
        st.markdown("##### Remove a Scheme")
        _del_options = [s["scheme_name"] for s in all_schemes]
        _del_c1, _del_c2 = st.columns([3, 1])
        with _del_c1:
            _del_scheme = st.selectbox(
                "Select scheme to remove",
                _del_options, key="del_scheme_select")
        with _del_c2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🗑️ Remove", key="del_scheme_btn"):
                delete_investment_scheme(_del_scheme)
                st.success(f"Removed {_del_scheme}")
                st.rerun()

    # ── Manual FD/MF Entry ────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Add Manual FD / MF Entry")
    st.caption("Add investments or redemptions not captured in bank statements.")

    with st.form("manual_inv_form", clear_on_submit=True):
        _mf1, _mf2, _mf3 = st.columns(3)
        with _mf1:
            _mi_date = st.date_input(
                "Date", value=datetime.date.today(),
                key="mi_date", format="DD/MM/YYYY")
            _mi_type = st.selectbox("Type (FD/MF)", ["FD", "MF"], key="mi_type")
        with _mf2:
            _mi_scheme = st.text_input(
                "Scheme Name",
                placeholder="e.g. SBI FD, HDFC Flexi Cap",
                key="mi_scheme")
            _mi_entity = st.selectbox(
                "Entity", ["Stores", "Ventures"], key="mi_entity")
        with _mf3:
            _mi_amount = st.number_input(
                "Amount (₹)", min_value=0.0, step=10000.0, key="mi_amount")
            _mi_entry_type = st.selectbox(
                "Invested / Redeemed", ["Invested", "Redeemed"],
                key="mi_entry_type")

        _mi_notes = st.text_input(
            "Notes (optional)",
            placeholder="Maturity date, interest rate etc.",
            key="mi_notes")

        if st.form_submit_button("Save Entry"):
            if not _mi_scheme.strip():
                st.error("Scheme Name is required.")
            elif _mi_amount <= 0:
                st.error("Amount must be greater than 0.")
            else:
                add_manual_investment(
                    entry_date  = str(_mi_date),
                    scheme_name = _mi_scheme.strip(),
                    scheme_type = _mi_type,
                    amount      = _mi_amount,
                    entry_type  = _mi_entry_type,
                    entity      = _mi_entity,
                    notes       = _mi_notes.strip(),
                )
                st.success(
                    f"{_mi_entry_type} {fmt_inr(_mi_amount)} in "
                    f"{_mi_scheme} ({_mi_type}) saved.")
                st.cache_data.clear()
                st.rerun()

    st.markdown("#### Manual Entry History")
    _manual_entries = get_manual_investments()
    if not _manual_entries:
        st.info("No manual entries yet.")
    else:
        for _me in _manual_entries:
            _me_color = "#16A34A" if _me["entry_type"] == "Invested" else "#DC2626"
            _me_cols  = st.columns([1.5, 2.5, 1, 1, 2, 1.5, 0.6])
            for _col, _val in zip(_me_cols[:6], [
                _me["entry_date"],
                _me["scheme_name"],
                _me["scheme_type"],
                _me["entity"],
                f'<b style="color:{_me_color};">'
                f'{_me["entry_type"]}: {fmt_inr(_me["amount"])}</b>',
                _me["notes"] or "—",
            ]):
                with _col:
                    st.markdown(
                        f'<div style="font-size:12px;padding:4px 0;">{_val}</div>',
                        unsafe_allow_html=True)
            with _me_cols[6]:
                if st.button("🗑️", key=f"del_mi_{_me['id']}",
                             help="Delete this entry"):
                    delete_manual_investment(_me["id"])
                    st.cache_data.clear()
                    st.rerun()

        st.markdown("---")
        st.markdown("#### Manual Entry Summary by Scheme")
        _mi_totals = get_manual_investment_totals()
        for _mt in _mi_totals:
            _mt_net   = (_mt["invested"] or 0) - (_mt["redeemed"] or 0)
            _mt_color = "#16A34A" if _mt_net >= 0 else "#DC2626"
            st.markdown(f"""
<div style="background:#FFFFFF;border-radius:8px;padding:12px 16px;
     border:1px solid #E8ECF0;margin:6px 0;
     display:flex;justify-content:space-between;align-items:center;">
    <div>
        <b style="font-size:13px;">{_mt['scheme_name']}</b>
        <span style="font-size:11px;color:#8896A5;margin-left:8px;">
            {_mt['scheme_type']} · {_mt['entity']}
        </span>
    </div>
    <div style="text-align:right;">
        <span style="font-size:12px;color:#4A5568;">
            Invested: {fmt_inr(_mt['invested'])} | Redeemed: {fmt_inr(_mt['redeemed'])}
        </span><br>
        <b style="font-size:13px;color:{_mt_color};">Net: {fmt_inr(_mt_net)}</b>
    </div>
</div>""", unsafe_allow_html=True)

    # ── B5: Transaction tagging ───────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Tag Investment Transactions")
    st.caption(
        "All transactions with Main Group = INVESTMENT. "
        "Add scheme name and number for portfolio tracking.")

    _inv_txns = get_investment_transactions(
        date_from=str(inv_from),
        date_to=str(inv_to),
        entity=entity_filter,
    )

    if not _inv_txns:
        st.info("No investment transactions found.")
    else:
        _df_it = pd.DataFrame(_inv_txns)
        _df_display = _df_it.copy()
        _df_display["debit"]  = _df_display["debit"].apply(
            lambda x: fmt_inr(x) if float(x or 0) > 0 else "—")
        _df_display["credit"] = _df_display["credit"].apply(
            lambda x: fmt_inr(x) if float(x or 0) > 0 else "—")
        _df_display.insert(0, "Tag", False)

        _edited = st.data_editor(
            _df_display[[
                "Tag", "id", "date", "entity", "bank",
                "narration", "debit", "credit",
                "final_group", "scheme_name",
                "scheme_number", "scheme_type"
            ]],
            column_config={
                "Tag":           st.column_config.CheckboxColumn("✓", width="small"),
                "id":            st.column_config.NumberColumn("ID", width="small"),
                "date":          st.column_config.Column("Date", width="small"),
                "narration":     st.column_config.Column("Narration", width="large"),
                "scheme_name":   st.column_config.Column("Scheme Name", width="medium"),
                "scheme_number": st.column_config.Column("Scheme No.", width="medium"),
                "scheme_type":   st.column_config.SelectboxColumn(
                    "Type", options=["FD", "MF", ""], width="small"),
            },
            disabled=["id", "date", "entity", "bank", "narration",
                      "debit", "credit", "final_group"],
            hide_index=True,
            use_container_width=True,
            height=400,
            key="inv_txn_editor",
        )

        _t1, _t2 = st.columns([2, 6])
        with _t1:
            if st.button("Save Tags", key="save_inv_tags"):
                _saved = 0
                for _, _trow in _edited[_edited["Tag"] == True].iterrows():
                    if str(_trow.get("scheme_name", "")).strip():
                        tag_investment_transaction(
                            transaction_id=int(_trow["id"]),
                            scheme_name=str(_trow["scheme_name"]).strip(),
                            scheme_number=str(_trow.get("scheme_number", "")).strip(),
                            scheme_type=str(_trow.get("scheme_type", "")).strip(),
                        )
                        _saved += 1
                if _saved:
                    st.success(f"Tagged {_saved} transactions")
                    st.cache_data.clear()
                    st.rerun()
                else:
                    st.warning("Select rows and add scheme names first.")

# ══════════════════════════════════════════════════════════════════════════════
# UPLOAD
# ══════════════════════════════════════════════════════════════════════════════
elif selected_tab == "Upload":
    import tempfile as _tempfile
    import sys as _sys
    import os as _os_up

    st.markdown("### Upload Bank Statements")
    st.markdown("""
<div style="background:#EBF4FF; border-left:4px solid #3B82F6;
     border-radius:6px; padding:10px 16px; margin-bottom:20px;
     font-size:13px; color:#1E40AF;">
    Select entity and account, then upload one bank statement file (.xlsx / .xls).
    Duplicate transactions are automatically skipped.
</div>
""", unsafe_allow_html=True)

    from config import ENTITIES as _UP_ENTITIES

    _up_c1, _up_c2 = st.columns(2)
    with _up_c1:
        _up_entity = st.selectbox(
            "Entity",
            list(_UP_ENTITIES.keys()),
            key="up_entity",
        )
    with _up_c2:
        _up_account = st.selectbox(
            "Account",
            _UP_ENTITIES[_up_entity],
            key="up_account",
        )

    # Uploader key rotation — allows auto-clear after processing
    if "uploader_key" not in st.session_state:
        st.session_state["uploader_key"] = 0
    if st.session_state.get("_clear_result_next"):
        st.session_state.pop("last_process_result", None)
        st.session_state.pop("_clear_result_next", None)

    _uploaded_file = st.file_uploader(
        "Bank Statement File",
        accept_multiple_files=False,
        type=["xlsx", "xls"],
        key=f"uploader_{st.session_state['uploader_key']}",
        help="AXIS: .xlsx   |   HDFC: .xls   |   One file at a time.",
    )
    if _uploaded_file is not None:
        st.session_state["_clear_result_next"] = True

    # Result banner (persists across reruns until dismissed)
    _up_result = st.session_state.get("last_process_result")
    if _up_result:
        if _up_result["status"] == "success":
            st.markdown(f"""
<div style="background:#F0FFF4; border-left:4px solid #38A169;
     border-radius:6px; padding:12px 16px; margin:8px 0; font-size:13px;">
    <b>{_up_result['filename']}</b> processed successfully<br>
    <span style="color:#276749;">
        {_up_result.get('entity','?')} — {_up_result.get('bank','?')} &nbsp;|&nbsp;
        {_up_result.get('inserted',0):,} rows inserted &nbsp;|&nbsp;
        {_up_result.get('skipped',0):,} duplicates skipped
    </span>
</div>""", unsafe_allow_html=True)
        elif _up_result["status"] == "already_loaded":
            st.markdown(f"""
<div style="background:#FFFFF0; border-left:4px solid #D69E2E;
     border-radius:6px; padding:12px 16px; margin:8px 0; font-size:13px;">
    <b>{_up_result['filename']}</b> — already loaded<br>
    <span style="color:#744210;">
        All {_up_result.get('skipped',0):,} rows already exist in the database.
    </span>
</div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
<div style="background:#FFF5F5; border-left:4px solid #E53E3E;
     border-radius:6px; padding:12px 16px; margin:8px 0; font-size:13px; color:#742A2A;">
    <b>{_up_result['filename']}</b> — {_up_result.get('message','')}
</div>""", unsafe_allow_html=True)
        if st.button("Dismiss", key="dismiss_result"):
            del st.session_state["last_process_result"]
            st.rerun()

    if st.button("Process Statement", key="up_process_btn",
                 disabled=not _uploaded_file, type="primary"):
        _script_dir = _os_up.path.dirname(_os_up.path.abspath(__file__))
        if _script_dir not in _sys.path:
            _sys.path.insert(0, _script_dir)
        from processor import process_file as _process_file
        from categorizer import load_keywords as _load_kw

        _suffix  = _os_up.path.splitext(_uploaded_file.name)[1].lower() or ".xlsx"
        _tmp_path = None
        try:
            with _tempfile.NamedTemporaryFile(
                    delete=False, suffix=_suffix, dir=_script_dir) as _tmp:
                _tmp.write(_uploaded_file.getbuffer())
                _tmp_path = _tmp.name

            with st.spinner(f"Processing {_uploaded_file.name}..."):
                _kw = _load_kw()
                _ins, _skp = _process_file(
                    _tmp_path,
                    keywords=_kw,
                    entity=_up_entity,
                    bank=_up_account,
                    delete_after=True,
                )

            if _ins > 0:
                # Derive date range and FY from inserted rows
                import sqlite3 as _sq3
                _conn2 = _sq3.connect(DATABASE_FILE)
                _dr = _conn2.execute("""
                    SELECT MIN(date), MAX(date), financial_year
                    FROM transactions
                    WHERE source_file=? AND entity=? AND bank=?
                    ORDER BY loaded_at DESC LIMIT 1
                """, [_uploaded_file.name, _up_entity, _up_account]).fetchone()
                _conn2.close()
                _d_from = _dr[0] if _dr else ""
                _d_to   = _dr[1] if _dr else ""
                _fy     = _dr[2] if _dr else ""
                log_upload(
                    filename=_uploaded_file.name,
                    entity=_up_entity,
                    bank=_up_account,
                    financial_year=_fy,
                    rows_inserted=_ins,
                    date_from=_d_from,
                    date_to=_d_to,
                )
                st.session_state["last_process_result"] = {
                    "status": "success", "filename": _uploaded_file.name,
                    "inserted": _ins, "skipped": _skp,
                    "entity": _up_entity, "bank": _up_account,
                }
            elif _skp > 0:
                st.session_state["last_process_result"] = {
                    "status": "already_loaded", "filename": _uploaded_file.name,
                    "skipped": _skp,
                }
            else:
                st.session_state["last_process_result"] = {
                    "status": "error", "filename": _uploaded_file.name,
                    "message": "No rows processed. Check file format.",
                }
        except Exception as _e:
            if _tmp_path and _os_up.path.exists(_tmp_path):
                try: _os_up.unlink(_tmp_path)
                except OSError: pass
            st.session_state["last_process_result"] = {
                "status": "error", "filename": _uploaded_file.name,
                "message": str(_e),
            }

        st.session_state["uploader_key"] += 1
        st.cache_data.clear()
        st.rerun()

    # ── DB Restore ────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### Restore Database")
    st.markdown("""
<div style="background:#FFF8E1; border-left:4px solid #F59E0B;
     border-radius:6px; padding:10px 16px; margin-bottom:16px;
     font-size:13px; color:#92400E;">
    ⚠️ Upload a <b>bankflow.db</b> file to restore all historical data.
    This will <b>replace</b> the current database entirely.
</div>
""", unsafe_allow_html=True)

    _db_file = st.file_uploader(
        "Select bankflow.db",
        type=["db"],
        key="db_restore_file",
        help="Upload your existing bankflow.db to restore all data."
    )

    _db_confirm = st.checkbox("I understand this will replace the current database", key="db_restore_confirm")
    _db_btn = st.button("🗄️ Restore Database", key="db_restore_btn",
                        disabled=not (_db_file and _db_confirm),
                        type="primary")

    if _db_btn and _db_file and _db_confirm:
        import shutil as _shutil
        from config import DATABASE_FILE as _DB_PATH
        try:
            _os_up.makedirs(_os_up.path.dirname(_DB_PATH), exist_ok=True)
            _backup_path = _DB_PATH + ".bak"
            if _os_up.path.exists(_DB_PATH):
                _shutil.copy2(_DB_PATH, _backup_path)
            with open(_DB_PATH, "wb") as _f:
                _f.write(_db_file.getbuffer())
            st.cache_data.clear()
            st.success(f"✅ Database restored successfully! ({_db_file.size:,} bytes written)")
            st.info("A backup of the previous database was saved as bankflow.db.bak")
        except Exception as _e:
            st.error(f"❌ Restore failed: {_e}")
